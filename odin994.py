import argparse
from PIL import Image
import cv2
import numpy as np
import os
import pandas as pd
import datetime
import tensorflow as tf
import sys
from utils import visualization_utils as vis_util
from object_detection.utils import ops as utils_ops
from utils import label_map_util
from collections import defaultdict
from io import StringIO
from matplotlib import pyplot as plt
from PIL import Image
from object_detection.utils import label_map_util
from object_detection.utils import visualization_utils as vis_util
import os
import PyPDF2
from fractions import Fraction
import unidecode
import math
import smtplib
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import smtplib
import yagmail
import math
import time
import shutil
import glob
import yagmail
import pathlib
import difflib
from google.cloud import vision
import io
from distutils.dir_util import copy_tree
from time import sleep
import re
import csv
from hermes2 import download_info
import json
import requests

tf.enable_eager_execution()


PATH_TO_CKPT = 'D:/models/research/object_detection/exported/competitive/frozen_inference_graph.pb'
PATH_TO_LABELS = 'D:/models/research/object_detection/exported/competitive/labelmap2.pbtxt'
PATH_TO_CKPT2 = 'D:/models/research/object_detection/exported/detection/frozen_inference_graph.pb'
PATH_TO_LABELS2 = 'D:/models/research/object_detection/exported/detection/labelmap.pbtxt'
NUM_CLASSES = 6
NUM_CLASSES2 = 7
PATH_TO_CKPT5 = 'D:/models/research/object_detection/exported/fencing/frozen_inference_graph.pb'
PATH_TO_LABELS5 = 'D:/models/research/object_detection/exported/fencing/labelmapfences.pbtxt'
NUM_CLASSES5 = 5

global namecount
namecount = int(0)
global sidingcount
sidingcount = int(0)
global fencecount
fencecount = int(0)


def get_cookies(a_dir):
    return [name for name in os.listdir(a_dir)
        if os.path.isdir(os.path.join(a_dir, name))]

def labor_analysis(zipcode):

    if len(zipcode) == 4:
        zipcode = "0"+zipcode
    if len(zipcode) == 3:
        zipcode = "00"+zipcode

    global pricelist
    global pricelist2

    oes = {
        "roofer": "47-2181",
        "roofer-jr": "47-3016",
        "siding": "47-4090",
        "mason": "47-2021",
        "siding": "47-2161",

        "drywaller": "47-2081",
        "insulation": "47-2131",
        "painter": "47-2141",
        "finish carpenter": "47-3012",
        "flooring": "47-2031",
        "wrought iron": "47-2171",
        "fence": "47-4031",
        "electrician": "47-2111",
        "electrician-jr": "47-3013",
        "carpet": "47-2041",
        "vinyl": "47-2042",
        "wood": "47-2031",
        "garbageman": "47-3019",
    }

    results_labor = {}
    with open('GFG.csv') as File:
        reader = csv.reader(File)
        for row in reader:
            if len(row)>5:
                state = row[1]
                oess = row[2]
                results_labor[state+'...'+oess]= row[13]

    results_zip = {}
    with open('uszips.csv') as File:
        reader = csv.reader(File)
        for row in reader:
            if len(row[0])==3:
                zip = '00'+row[0]
            elif len(row[0])==4:
                zip = '0'+row[0]
            else:
                zip = row[0]
            results_zip[zip] = row[3],row[4]

    state = results_zip.get(zipcode)[0]
    pop_density = float(results_zip.get(zipcode)[1])
    labor_local_factor = float(1)
    if pop_density > float(10000):
        labor_local_factor +=float(.275)
    elif pop_density > float(4000):
        labor_local_factor +=float(.20)
    elif pop_density > float(2500):
        labor_local_factor +=float(.125)
    elif pop_density > float(750):
        labor_local_factor +=float(.065)
    elif pop_density < float(300):
        labor_local_factor += float(-.05)
    elif pop_density < float(300):
        labor_local_factor +=float(-.075)

    trades = oes.values()
    local_labor_list = []

    labor_burden = 1.35

    for trade in trades:
        get_string = state+'...'+trade
        raw_state_labor = results_labor.get(get_string)
        local_single_labor = float(raw_state_labor)*labor_local_factor*float(1+labor_burden)
        local_labor_list.append(local_single_labor)

    keys_list = oes.keys()
    labor = {}

    r = 0
    for each in keys_list:
        labor[each]=local_labor_list[r]
        r +=1

    pricelist2 = {

      "Board and Batten": [.05*labor.get('siding'),5],
      "Masonry Siding": [.025*labor.get('mason'),3.50],
      "Lap Siding": [.0425*labor.get("siding"),1.75],
      "Shingled Siding": [.055*labor.get("siding"), 3.80],
      "Stone Siding": [.05*labor.get("mason"), 9.55],
      "Stucco Siding": [.1*labor.get("siding"), 2],
      "Log Siding": [.04*labor.get("siding"), 5.90],
      "House Wrap": [.02*labor.get("siding"), .2],
      "Corner Trim": [.05*labor.get("siding"), 3.50],
      "Window Trim": [1*labor.get("siding"), 25.00],

      "Laminated Shingles": [1.4*labor.get("roofer"), 120],
      "3-Tab Shingles": [1.3*labor.get("roofer"), 90],
      "Roof Tile Demo": [2.75*labor.get("roofer-jr"), 0],
      "S-Shaped Clay Tiles": [2.75*labor.get("roofer"), 340],
      "Built Up Roofing": [2.4*labor.get("roofer"), 93.51],
      "Shingle Removal": [.85*labor.get("roofer-jr"), 0],
      "Metal Roofing": [3.25*labor.get('roofer'), 140],
      "Metal Roofing Removal": [1*labor.get('roofer-jr'), 0],
      "Metal Ridge Cap": [.045*labor.get('roofer'), 2.43],
      "Roof Felt": [.2*labor.get("roofer"), 6],
      "Built Up Removal": [.85*labor.get('roofer-jr'), 0],
      "Flat Tile": [2.85*labor.get("roofer"), 350],
      "Ridge Cap": [.03*labor.get('roofer'), 1.13],
      "Tile Ridge Cap": [.04*labor.get('roofer'), 4.57],
      "Metal Ridge Cap": [.045*labor.get('roofer'), 2.65],
      "Roof Vent Flashing": [.5*labor.get('roofer'), 12.23],
      "Drip Edge": [.02*labor.get('roofer'), .82],
      "Detach and Reset Solar Panel": [2.5*labor.get('electrician'), 0],

      "Steep Charge": [.3*labor.get('roofer'), 0],
      "Very Steep Charge": [.8*labor.get('roofer'), 0],
      "High Charge": [.25*labor.get('roofer'), 0],

      "Valley Metal": [.1*labor.get('roofer'), 10.13],
      "Vented Ridge Cap": [.075*labor.get('roofer'), 3],
      "Chimney Flashing": [2.5*labor.get('roofer'), 80],
      "Detach Satellite Dish": [.8*labor.get('roofer'), 0],
      "Skylight Flashing": [.4*labor.get('roofer'), 88.75],

      "Drywall": [.0325*labor.get("drywaller"), .52],
      "Paint": [.01*labor.get("painter"), .25],
      "Baseboards": [.031*labor.get('finish carpenter'),.85],
      "Paint Baseboards": [.025*labor.get('painter'), .18],
      "Insulation": [.02*labor.get('drywaller'), .8],

      "Board Fence": [.35*labor.get('fence'), 22],
      "Chain Link Fence": [.225*labor.get('fence'), 7],
      "Wrought Iron Fence": [.3*labor.get('fence'), 45],
      "Picket Fence": [.02*labor.get('fence'), 7.84],
      "Rail Fence": [.25*labor.get("fence"), 10.75],
      "Paint Fence": [.0125*labor.get("painter"),.29],

      "Carpet": [.0225*labor.get('flooring'), 3],
      "Vinyl": [.0425*labor.get('flooring'), 2],
      "Engineered": [.1*labor.get('flooring'), 7.25],
      "Laminated": [.06*labor.get('flooring'), 3.25],
      "Vapor Barrier": [.2*labor.get("flooring"), .45],
      "Reducer Strip": [.015*labor.get("flooring"), 6.05],

      "Debris Haul": [2.5*labor.get("garbageman"), 60],
      "Dumpster": [0, 575],
    }

    pricelist = {

      "Board and Batten": round((sum(pricelist2.get('Board and Batten'))),2),
      "Masonry Siding": round((sum(pricelist2.get('Masonry Siding'))),2),
      "Lap Siding": round((sum(pricelist2.get('Lap Siding'))),2),
      "Shingled Siding": round((sum(pricelist2.get('Shingled Siding'))),2),
      "Stone Siding": round((sum(pricelist2.get('Stone Siding'))),2),
      "Stucco Siding": round((sum(pricelist2.get('Stucco Siding'))),2),
      "Log Siding":round((sum(pricelist2.get('Log Siding'))),2),
      "House Wrap": round((sum(pricelist2.get("House Wrap"))),2),
      "Corner Trim": round((sum(pricelist2.get("Corner Trim"))),2),
      "Window Trim": round((sum(pricelist2.get("Window Trim"))),2),

      "Laminated Shingles": round((sum(pricelist2.get('Laminated Shingles'))),2),
      "3-Tab Shingles": round((sum(pricelist2.get('3-Tab Shingles'))),2),
      "Roof Tile Demo": round((sum(pricelist2.get('Roof Tile Demo'))),2),
      "S-Shaped Clay Tiles": round((sum(pricelist2.get('S-Shaped Clay Tiles'))),2),
      "Built Up Roofing": round((sum(pricelist2.get('Built Up Roofing'))),2),
      "Shingle Removal": round(sum(pricelist2.get("Shingle Removal")),2),
      "Metal Roofing": round((sum(pricelist2.get('Metal Roofing'))),2),
      "Metal Roofing Removal": round((sum(pricelist2.get('Metal Roofing Removal'))),2),
      "Metal Ridge Cap": round((sum(pricelist2.get('Metal Ridge Cap'))),2),
      "Roof Felt": round((sum(pricelist2.get('Roof Felt'))),2),
      "Built Up Removal": round((sum(pricelist2.get('Built Up Removal'))),2),
      "Flat Tile": round((sum(pricelist2.get('Flat Tile'))),2),
      "Ridge Cap": round((sum(pricelist2.get('Ridge Cap'))),2),
      "Tile Ridge Cap": round((sum(pricelist2.get('Tile Ridge Cap'))),2),
      "Metal Ridge Cap": round((sum(pricelist2.get('Metal Ridge Cap'))),2),
      "Roof Vent Flashing": round((sum(pricelist2.get('Roof Vent Flashing'))),2),
      "Drip Edge": round((sum(pricelist2.get("Drip Edge"))),2),
      "Detach and Reset Solar Panel": round((sum(pricelist2.get("Detach and Reset Solar Panel"))),2),

      "Steep Charge": round((sum(pricelist2.get('Steep Charge'))),2),
      "Very Steep Charge": round((sum(pricelist2.get("Very Steep Charge"))),2),
      "High Charge": round((sum(pricelist2.get('High Charge'))),2),

      "Valley Metal": round((sum(pricelist2.get('Valley Metal'))),2),
      "Vented Ridge Cap": round((sum(pricelist2.get('Vented Ridge Cap'))),2),
      "Chimney Flashing": round((sum(pricelist2.get('Chimney Flashing'))),2),
      "Detach Satellite Dish": round((sum(pricelist2.get("Detach Satellite Dish"))),2),
      "Skylight Flashing": round((sum(pricelist2.get('Skylight Flashing'))),2),

      "Drywall": round((sum(pricelist2.get('Drywall'))),2),
      "Paint": round((sum(pricelist2.get('Paint'))),2),
      "Baseboards": round((sum(pricelist2.get("Baseboards"))),2),
      "Paint Baseboards": round((sum(pricelist2.get("Paint Baseboards"))),2),
      "Insulation": round((sum(pricelist2.get("Insulation"))),2),

      "Board Fence": round((sum(pricelist2.get("Board Fence"))),2),
      "Chain Link Fence": round((sum(pricelist2.get("Chain Link Fence"))),2),
      "Wrought Iron Fence": round((sum(pricelist2.get("Wrought Iron Fence"))),2),
      "Picket Fence": round((sum(pricelist2.get("Picket Fence"))),2),
      "Rail Fence": round((sum(pricelist2.get("Rail Fence"))),2),
      "Paint Fence": round((sum(pricelist2.get("Paint Fence"))),2),

      "Carpet": round((sum(pricelist2.get("Carpet"))),2),
      "Vinyl": round((sum(pricelist2.get("Vinyl"))),2),
      "Engineered": round((sum(pricelist2.get("Engineered"))),2),
      "Laminated": round((sum(pricelist2.get("Laminated"))),2),
      "Reducer Strip": round((sum(pricelist2.get('Reducer Strip'))),2),

      "Debris Haul": round((sum(pricelist2.get("Debris Haul"))),2),
      "Dumpster": round((sum(pricelist2.get("Dumpster"))),2),
    }

    return pricelist
    return pricelist2


def get_admin(batch):
        wb = openpyxl.load_workbook('D:/models/research/object_detection/template.xlsx')

        #insert logo


        x = batch
        PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
        f=open(PATH_TO_TEST_INFO, "r")
        contents =f.read()
        lines = contents.splitlines()
        clientname = lines[1][:-3]
        if len(clientname) <= 1:
            clientname = x
        clientemail = lines[2][:-3]
        clientphone = lines[3][:-3]
        recipientemail = lines[4][:-3].strip()
        latlong = lines[5][:-3]
        clientaddress = lines[6][:-3]

        if len(lines) >= 8:
            replace = lines[7][:-3]
        if len(lines) >= 9:
            repaint = lines[8][:-3]

        sheet = wb.get_sheet_by_name('Roofing')
        sheet2 = wb.get_sheet_by_name('Siding')
        sheet3 = wb.get_sheet_by_name('Fencing')
        sheet4 = wb.get_sheet_by_name('Interior')
        sheet5 = wb.get_sheet_by_name('PriceList')
        sheet6 = wb.get_sheet_by_name('Summary')
        sheet7 = wb.get_sheet_by_name('Photos')
        sheet['B8'] = clientname
        sheet['B9'] = clientaddress
        sheet['E8'] = clientphone
        sheet['E9'] = clientemail
        sheet2['B8'] = clientname
        sheet2['B9'] = clientaddress
        sheet2['E8'] = clientphone
        sheet2['E9'] = clientemail
        sheet3['B8'] = clientname
        sheet3['B9'] = clientaddress
        sheet3['E8'] = clientphone
        sheet3['E9'] = clientemail
        sheet4['B8'] = clientname
        sheet4['B9'] = clientaddress
        sheet4['E8'] = clientphone
        sheet4['E9'] = clientemail
        sheet6['B8'] = clientname
        sheet6['B9'] = clientaddress
        sheet6['E8'] = clientphone
        sheet6['E9'] = clientemail
        sheet7['B8'] = clientname
        sheet7['B9'] = clientaddress
        sheet7['E8'] = clientphone
        sheet7['E9'] = clientemail


        img = openpyxl.drawing.image.Image('D:/models/research/object_detection/logo.jpg')
        img2 = openpyxl.drawing.image.Image('D:/models/research/object_detection/logo.jpg')
        img3 = openpyxl.drawing.image.Image('D:/models/research/object_detection/logo.jpg')
        img4 = openpyxl.drawing.image.Image('D:/models/research/object_detection/logo.jpg')
        img6 = openpyxl.drawing.image.Image('D:/models/research/object_detection/logo.jpg')
        img7 = openpyxl.drawing.image.Image('D:/models/research/object_detection/logo.jpg')
        img.anchor = 'A1'
        img2.anchor = 'A1'
        img3.anchor = 'A1'
        img4.anchor = 'A1'
        img6.anchor = 'A1'
        img7.anchor = 'A1'
        sheet.add_image(img)
        sheet2.add_image(img2)
        sheet3.add_image(img3)
        sheet4.add_image(img4)
        sheet6.add_image(img6)
        sheet7.add_image(img7)
        zip = '60565'
        zip = clientaddress[-5:].strip()
        if zip.isnumeric() == True and len(zip)<=5:
            labor_analysis(zip)
        else:
            labor_analysis('60565')
        csv_file = csv.reader(open('zips.csv', "r"), delimiter=",")
        for row in csv_file:
            if zip == row[0]:
                taxrate = row[1]


        else:
            taxrate=.07

        for row, (key, price) in enumerate(pricelist.items(), start=2):
            sheet5 [f"A{row}"] = key
            sheet5 [f"B{row}"] = price
        wb.save('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')


PATH_TO_CKPT6 = 'D:/models/research/object_detection/exported/story2/frozen_inference_graph.pb'
PATH_TO_LABELS6 = 'D:/models/research/object_detection/exported/story2/labelmap.pbtxt'
NUM_CLASSES6 = 9

def slopey(batch):
    zz = 0
    counted = 0
    slopelist = [0,0,0,0,0]
    levellist = [0,0]

    roofingraw = 0
    x = batch
    for path in pathlib.Path('D:/models/research/object_detection/server/connect/'+x+'/roofing/').iterdir():
        if path.is_file():
            roofingraw = roofingraw+1


    if roofingraw >= 1:
        PATH_TO_TEST_IMAGES_DIR = 'D:/models/research/object_detection/server/connect/'+x+'/roofing/'
        PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
                    #rename the photos
        def rename():
            n = 1
            for name in os.listdir(PATH_TO_TEST_IMAGES_DIR):
                dst = PATH_TO_TEST_IMAGES_DIR + "image" + str(n) + ".jpg"
                src = PATH_TO_TEST_IMAGES_DIR + name
                os.rename(src, dst)
                n = n+1
        rename()

        f=open(PATH_TO_TEST_INFO, "r")

        roofingraw+=1
        TEST_IMAGE_PATHS = [ os.path.join(PATH_TO_TEST_IMAGES_DIR, 'image{}.jpg'.format(i)) for i in range(1, roofingraw) ]

        #simple counters

        features = [0,0,0,0,0,0,0,0,0,0]
        photolist = [""]

        #set up inference
        def run_inference_for_single_image(image, sess, tensor_dict):
            if 'detection_masks' in tensor_dict:

                # The following processing is only for single image
                detection_boxes = tf.squeeze(tensor_dict['detection_boxes'], [0])
                detection_masks = tf.squeeze(tensor_dict['detection_masks'], [0])
                # Reframe is required to translate mask from box coordinates to image coordinates and fit the image size.
                real_num_detection = tf.cast(tensor_dict['num_detections'][0], tf.int32)
                detection_boxes = tf.slice(detection_boxes, [0, 0], [real_num_detection, -1])
                detection_masks = tf.slice(detection_masks, [0, 0, 0], [real_num_detection, -1, -1])
                detection_masks_reframed = utils_ops.reframe_box_masks_to_image_masks(
                    detection_masks, detection_boxes, image.shape[0], image.shape[1])
                detection_masks_reframed = tf.cast(
                    tf.greater(detection_masks_reframed, 0.5), tf.uint8)
                # Follow the convention by adding back the batch dimension
                tensor_dict['detection_masks'] = tf.expand_dims(detection_masks_reframed, 0)
            image_tensor = tf.get_default_graph().get_tensor_by_name('image_tensor:0')

            # Run inference
            output_dict = sess.run(tensor_dict, feed_dict={image_tensor: np.expand_dims(image, 0)})

            # all outputs are float32 numpy arrays, so convert types as appropriate
            output_dict['num_detections'] = int(output_dict['num_detections'][0])
            output_dict['detection_classes'] = output_dict[
                'detection_classes'][0].astype(np.uint8)
            output_dict['detection_boxes'] = output_dict['detection_boxes'][0]
            output_dict['detection_scores'] = output_dict['detection_scores'][0]
            if 'detection_masks' in output_dict:
                output_dict['detection_masks'] = output_dict['detection_masks'][0]
            return output_dict

        detection_graph = tf.Graph()
        with detection_graph.as_default():
          od_graph_def = tf.GraphDef()
          with tf.gfile.GFile(PATH_TO_CKPT6, 'rb') as fid:
            serialized_graph = fid.read()
            od_graph_def.ParseFromString(serialized_graph)
            tf.import_graph_def(od_graph_def, name='')

        label_map = label_map_util.load_labelmap(PATH_TO_LABELS6)
        categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES6, use_display_name=True)
        category_index = label_map_util.create_category_index(categories)

        def load_image_into_numpy_array(image):
          (im_width, im_height) = image.size
          return np.array(image.getdata()).reshape(
              (im_height, im_width, 3)).astype(np.uint8)



        with detection_graph.as_default():
          with tf.Session(graph=detection_graph) as sess:
            tf.enable_eager_execution()
            sess.run(tf.global_variables_initializer())
            tf.enable_eager_execution()
            ops = tf.get_default_graph().get_operations()
            tf.enable_eager_execution()
            all_tensor_names = {output.name for op in ops for output in op.outputs}
            tensor_dict = {}
            for key in [
                'num_detections', 'detection_boxes', 'detection_scores',
                'detection_classes', 'detection_masks'
            ]:
                tensor_name = key + ':0'
                if tensor_name in all_tensor_names:
                    tensor_dict[key] = tf.get_default_graph().get_tensor_by_name(tensor_name)
            img = 1
            for image_path in TEST_IMAGE_PATHS:
              tf.enable_eager_execution()
              image = Image.open(image_path)
              image_np = load_image_into_numpy_array(image)
              # Expand dimensions since the model expects images to have shape: [1, None, None, 3]
              image_np_expanded = np.expand_dims(image_np, axis=0)
              image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')
              # Each box represents a part of the image where a particular object was detected.
              boxes = detection_graph.get_tensor_by_name('detection_boxes:0')
              scores = detection_graph.get_tensor_by_name('detection_scores:0')
              classes = detection_graph.get_tensor_by_name('detection_classes:0')
              num_detections = detection_graph.get_tensor_by_name('num_detections:0')
              output_dict = run_inference_for_single_image(image_np, sess, tensor_dict)
              dictionary = output_dict['detection_scores']
              best_score = dictionary.item(0)
              second_best_score = dictionary.item(1)
              dictionarys = output_dict['detection_classes']
              best_label = dictionarys.item(0)
              second_best_label = dictionarys.item(1)
              detectedboxes = output_dict['detection_boxes']
              box1 = detectedboxes[0]
              box2 = detectedboxes[1]
              zz+=1

              foundlist = list(dictionarys)
              scorelist = list(dictionary)


              if 2 in foundlist or 3 in foundlist or 4 in foundlist or 5 in foundlist or 6 in foundlist:
                  counted+=1
              if 2 in foundlist:
                  index = foundlist.index(2)
                  slopelist[0] += scorelist[index]
              if 3 in foundlist:
                  index = foundlist.index(3)
                  slopelist[1] += scorelist[index]
              if 4 in foundlist:
                  index = foundlist.index(4)
                  slopelist[2] += scorelist[index]
              if 5 in foundlist:
                  index = foundlist.index(5)
                  slopelist[3] += scorelist[index]
              if 6 in foundlist:
                  index = foundlist.index(6)
                  slopelist[4] += scorelist[index]
              if 7 in foundlist:
                  index = foundlist.index(7)
                  levellist[0] += scorelist[index]
              if 8 in foundlist:
                  index = foundlist.index(8)
                  levellist[1]+= scorelist[index]


    if sum(slopelist) > 0:
        averageslope = [x / counted for x in slopelist]
        if slopelist[1]==0 and slopelist[2]==0 and slopelist[3]==0 and slopelist[4]==0 and slopelist[0] >= .5:
            slope = 0
        if averageslope[0] == 0 and slopelist[1]==0 and slopelist[3]==0 and slopelist[4]==0 and slopelist[3] >= .8:
            slope = 4
        else:
            slope = 5.30133-.873*averageslope[1]-1.85*averageslope[2]+7.87*averageslope[3]+9.25*averageslope[4]
        slope=round(slope)
        if levellist[0] > levellist[1]:
            levels = 1
        elif levellist[1] > levellist[0]:
            levels = 2
        else:
            levels = 1
        f=open(PATH_TO_TEST_INFO, "a")
        f.write('\nslope '+str(slope))
        f.write('\nlevels '+str(levels))
        f.close()


def measurements(batch):
        x= batch
        PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
        f=open(PATH_TO_TEST_INFO, "r")
        contents =f.read()
        lines = contents.splitlines()
        clientname = lines[1][:-3]
        if len(clientname) <= 1:
            clientname = x
        clientemail = lines[2][:-3]
        clientphone = lines[3][:-3]
        recipientemail = lines[4][:-3].strip()
        latlong = lines[5][:-3]
        clientaddress = lines[6][:-3]

        print("This is for "+clientaddress)
        elements = float(input("Number of roofing elements? "))
        ll = 1
        measurestring = ""

        while ll <= elements:
            sublist=[0,0,0]
            length = 0
            width = 0
            overlap = 0
            shape = float(input('What is the shape of element '+str(ll)+'? 1 = square or rectangle, 2 = hexagonal, 3 = circular, 4 = irregular '))
            if shape == 1:
                length = float(input("What is the length of the long side of element "+str(ll)+"? "))
                width = float(input("What is the width of the short side of element "+str(ll)+"? "))
            else:
                length = float(input("What is the length of the long side of element "+str(ll)+"? "))
                width = float(input("What is the width of the short side of element "+str(ll)+"? "))

            if ll>=2:
                overlap = float(input('What is the length of overlap?'))
            measurestring += str(length)+","+str(width)+","+str(overlap)+","
            ll+=1

        f=open(PATH_TO_TEST_INFO, "a")
        f.write('\nmeasure '+str(measurestring))
        f.close()



def roofing(batch):

        x = batch
        roofingraw = 0
        for path in pathlib.Path('D:/models/research/object_detection/server/connect/'+x+'/roofing/').iterdir():
            if path.is_file():
                roofingraw = roofingraw+1


        if roofingraw >= 1:
            PATH_TO_TEST_IMAGES_DIR = 'D:/models/research/object_detection/server/connect/'+x+'/roofing/'
            PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
                        #rename the photos
            def rename():
                n = 1
                for name in os.listdir(PATH_TO_TEST_IMAGES_DIR):
                    dst = PATH_TO_TEST_IMAGES_DIR + "image" + str(n) + ".jpg"
                    src = PATH_TO_TEST_IMAGES_DIR + name
                    os.rename(src, dst)
                    n = n+1
            rename()

            f=open(PATH_TO_TEST_INFO, "r")
            contents =f.read()
            lines = contents.splitlines()
            clientname = lines[1][:3]
            if len(clientname) <= 1:
                clientname = x
            clientaddress = lines[6][:-3]
            zip = clientaddress[-5:].strip()
            measurelist = []

            PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
            f=open(PATH_TO_TEST_INFO, "r")
            contents =f.read()
            lines = contents.splitlines()
            slope = 4
            levels = 1
            for line in lines:
                if line.startswith('measure'):
                    rawMeasure = line[(len('measure ')):]
                    measurelist = rawMeasure.split(',')
                    for each in measurelist:
                        if not each == "":
                            each = float(each)
                    measurelist = measurelist[:-1]

                if line.startswith('slope'):
                    slope = line[len('slope '):]
                    print('slope is '+slope)
                if line.startswith('levels'):
                    levels = line[7]

                    print('levelsraw are '+levels)

            if zip.isnumeric()==True and len(zip)==5 :
                labor_analysis(zip)
                csv_file = csv.reader(open('zips.csv', "r"), delimiter=",")


                #loop through the csv list
                for row in csv_file:
                    #if current rows 2nd value is equal to input, print that row
                    if zip == row[0]:
                        taxrate = row[1]


            else:

                taxrate=.07

            f.close()
            roofingraw+=1
            TEST_IMAGE_PATHS = [ os.path.join(PATH_TO_TEST_IMAGES_DIR, 'image{}.jpg'.format(i)) for i in range(1, roofingraw) ]

            #simple counters
            laminated = 0
            shingle = 0
            stile = 0
            seam = 0
            ftile = 0
            bit = 0
            gable = 0
            hip = 0
            chimney = 0
            vented = 0
            sat = 0
            skylight = 0
            solar = 0
            i = 0
            d = 0
            b = 0
            chimneycount = 0
            ventedridge = 0
            satcount = 0
            skylightcount = 0
            solarcount = 0
            shingletype = 0
            features = [0,0,0,0,0,0,0,0]
            photolist = [""]

            #set up inference
            def run_inference_for_single_image(image, sess, tensor_dict):
                if 'detection_masks' in tensor_dict:

                    # The following processing is only for single image
                    detection_boxes = tf.squeeze(tensor_dict['detection_boxes'], [0])
                    detection_masks = tf.squeeze(tensor_dict['detection_masks'], [0])
                    # Reframe is required to translate mask from box coordinates to image coordinates and fit the image size.
                    real_num_detection = tf.cast(tensor_dict['num_detections'][0], tf.int32)
                    detection_boxes = tf.slice(detection_boxes, [0, 0], [real_num_detection, -1])
                    detection_masks = tf.slice(detection_masks, [0, 0, 0], [real_num_detection, -1, -1])
                    detection_masks_reframed = utils_ops.reframe_box_masks_to_image_masks(
                        detection_masks, detection_boxes, image.shape[0], image.shape[1])
                    detection_masks_reframed = tf.cast(
                        tf.greater(detection_masks_reframed, 0.5), tf.uint8)
                    # Follow the convention by adding back the batch dimension
                    tensor_dict['detection_masks'] = tf.expand_dims(detection_masks_reframed, 0)
                image_tensor = tf.get_default_graph().get_tensor_by_name('image_tensor:0')

                # Run inference
                output_dict = sess.run(tensor_dict, feed_dict={image_tensor: np.expand_dims(image, 0)})

                # all outputs are float32 numpy arrays, so convert types as appropriate
                output_dict['num_detections'] = int(output_dict['num_detections'][0])
                output_dict['detection_classes'] = output_dict[
                    'detection_classes'][0].astype(np.uint8)
                output_dict['detection_boxes'] = output_dict['detection_boxes'][0]
                output_dict['detection_scores'] = output_dict['detection_scores'][0]
                if 'detection_masks' in output_dict:
                    output_dict['detection_masks'] = output_dict['detection_masks'][0]
                return output_dict

            detection_graph = tf.Graph()
            with detection_graph.as_default():
              od_graph_def = tf.GraphDef()
              with tf.gfile.GFile(PATH_TO_CKPT, 'rb') as fid:
                serialized_graph = fid.read()
                od_graph_def.ParseFromString(serialized_graph)
                tf.import_graph_def(od_graph_def, name='')

            label_map = label_map_util.load_labelmap(PATH_TO_LABELS)
            categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES, use_display_name=True)
            category_index = label_map_util.create_category_index(categories)

            def load_image_into_numpy_array(image):
              (im_width, im_height) = image.size
              return np.array(image.getdata()).reshape(
                  (im_height, im_width, 3)).astype(np.uint8)



            with detection_graph.as_default():
              with tf.Session(graph=detection_graph) as sess:
                tf.enable_eager_execution()
                sess.run(tf.global_variables_initializer())
                tf.enable_eager_execution()
                ops = tf.get_default_graph().get_operations()
                tf.enable_eager_execution()
                all_tensor_names = {output.name for op in ops for output in op.outputs}
                tensor_dict = {}
                for key in [
                    'num_detections', 'detection_boxes', 'detection_scores',
                    'detection_classes', 'detection_masks'
                ]:
                    tensor_name = key + ':0'
                    if tensor_name in all_tensor_names:
                        tensor_dict[key] = tf.get_default_graph().get_tensor_by_name(tensor_name)
                img = 1
                for image_path in TEST_IMAGE_PATHS:
                  tf.enable_eager_execution()
                  image = Image.open(image_path)
                  image_np = load_image_into_numpy_array(image)
                  # Expand dimensions since the model expects images to have shape: [1, None, None, 3]
                  image_np_expanded = np.expand_dims(image_np, axis=0)
                  image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')
                  # Each box represents a part of the image where a particular object was detected.
                  boxes = detection_graph.get_tensor_by_name('detection_boxes:0')
                  scores = detection_graph.get_tensor_by_name('detection_scores:0')
                  classes = detection_graph.get_tensor_by_name('detection_classes:0')
                  num_detections = detection_graph.get_tensor_by_name('num_detections:0')
                  output_dict = run_inference_for_single_image(image_np, sess, tensor_dict)
                  dictionary = output_dict['detection_scores']
                  best_score = dictionary.item(0)
                  second_best_score = dictionary.item(1)
                  dictionarys = output_dict['detection_classes']
                  best_label = dictionarys.item(0)
                  second_best_label = dictionarys.item(1)
                  detectedboxes = output_dict['detection_boxes']
                  box1 = detectedboxes[0]
                  box2 = detectedboxes[1]


                  #analysis of tesnsorsss
                  #throw out ambiguous
                  if best_score <= .9:
                    print('This photo yielded nothing useful')
                  if best_score >= .9 and second_best_score >= .9:
                    area1 = float(((box1[2]-box1[0])*(box1[3]-box1[1])))
                    area2 = float(((box2[2]-box2[0])*(box2[3]-box2[1])))
                    if area1 >= (float(2)*area2):
                        if best_label == 1:
                            laminated = laminated+1
                        if best_label == 2:
                            shingle=shingle+1
                        if best_label == 3:
                            stile = stile+1
                        if best_label == 4:
                            seam = seam+1
                        if best_label == 5:
                            ftile = ftile+1
                        if best_label == 6:
                            bit = bit+1
                    if area2 >= (float(2)*area1):
                        if second_best_label == 1:
                            laminated = laminated+1
                        if second_best_label == 2:
                            shingle=shingle+1
                        if second_best_label == 3:
                            stile = stile+1
                        if second_best_label == 4:
                            seam = seam+1
                        if second_best_label == 5:
                            ftile = ftile+1
                        if second_best_label == 6:
                            bit = bit+1
                    else:
                        azd = 0

                  if best_label == 1 and best_score >= .90:
                    laminated = laminated + 1

                  if best_label == 2 and best_score >= .90:
                    shingle = shingle + 1

                  if best_label == 3 and best_score >= .90:
                    stile = stile + 1

                  if best_label == 4 and best_score >= .90:
                    seam = seam + 1

                  if best_label == 5 and best_score >= .90:
                    ftile = ftile + 1

                  if best_label == 6 and best_score >= .95:
                    bit = bit + 1

            roofMaterials = float(0)
            rooflist = [0, laminated, shingle, stile, seam, ftile, bit]  #0 is a dummy to ensure calls to right values in other areas
            max_count = rooflist.index(max(rooflist))
            #if max(rooflist) < 2:
            #    max_count = 7
            if max_count == 1:
                rooftype = 1
                print('The roof was determinated to be laminated')
                photolist[0]=['laminated']
            elif max_count == 2:
                rooftype = 2
                print('The roof type was determined to be standard shingle')
                photolist[0]=['shingle']
            elif max_count == 3:
                rooftype = 3
                print('The roof type was determined to be s-tile')
                photolist[0]='s-tile'
            elif max_count == 4:
                rooftype = 4
                print('The roof type was determinated to be standing-seam')
                photolist[0]='metal'
            elif max_count == 5:
                rooftype = 5
                print('The roof type was determined to be flat tile')
                photolist[0]='flat'
            elif max_count == 6:
                rooftype = 6
                print('The roof type was determined to be bit roof')
                photolist[0]='bit'
            else:
                rooftype = 7
                print('the roof type was indeterminate')

            #run inference for detection objects

            detection_graph = tf.Graph()
            with detection_graph.as_default():
              od_graph_def = tf.GraphDef()
              with tf.gfile.GFile(PATH_TO_CKPT2, 'rb') as fid:
                serialized_graph = fid.read()
                od_graph_def.ParseFromString(serialized_graph)
                tf.import_graph_def(od_graph_def, name='')

            label_map = label_map_util.load_labelmap(PATH_TO_LABELS2)
            categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES2, use_display_name=True)
            category_index = label_map_util.create_category_index(categories)

            def load_image_into_numpy_array(image):
              (im_width, im_height) = image.size
              return np.array(image.getdata()).reshape(
                  (im_height, im_width, 3)).astype(np.uint8)



            with detection_graph.as_default():
              with tf.Session(graph=detection_graph) as sess:
                tf.enable_eager_execution()
                sess.run(tf.global_variables_initializer())
                tf.enable_eager_execution()
                ops = tf.get_default_graph().get_operations()
                tf.enable_eager_execution()
                all_tensor_names = {output.name for op in ops for output in op.outputs}
                tensor_dict = {}
                for key in [
                    'num_detections', 'detection_boxes', 'detection_scores',
                    'detection_classes', 'detection_masks'
                ]:
                    tensor_name = key + ':0'
                    if tensor_name in all_tensor_names:
                        tensor_dict[key] = tf.get_default_graph().get_tensor_by_name(tensor_name)
                img = 1
                imgs = 1
                for image_path in TEST_IMAGE_PATHS:
                  tf.enable_eager_execution()
                  image = Image.open(image_path)
                  image_np = load_image_into_numpy_array(image)
                  # Expand dimensions since the model expects images to have shape: [1, None, None, 3]
                  image_np_expanded = np.expand_dims(image_np, axis=0)
                  image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')
                  # Each box represents a part of the image where a particular object was detected.
                  boxes = detection_graph.get_tensor_by_name('detection_boxes:0')
                  scores = detection_graph.get_tensor_by_name('detection_scores:0')
                  classes = detection_graph.get_tensor_by_name('detection_classes:0')
                  num_detections = detection_graph.get_tensor_by_name('num_detections:0')
                  output_dict = run_inference_for_single_image(image_np, sess, tensor_dict)
                  detectedscores = output_dict['detection_scores']
                  detectedclasses = output_dict['detection_classes']


                  eligible1 = detectedscores[0]
                  eligible2 = detectedscores[1]
                  eligible3 = detectedscores[2]
                  eligible4 = detectedscores[3]
                  eligible5 = detectedscores[4]
                  eligibleclass1 =detectedclasses[0]
                  eligibleclass2 = detectedclasses[1]
                  eligibleclass3 = detectedclasses[2]
                  eligibleclass4 = detectedclasses[3]
                  eligibleclass5 = detectedclasses[4]


                 #since this is detection not competition, using lower threshold

                  detectionscoreslist = [eligible1, eligible2, eligible3, eligible4, eligible5]

                  detectedthreshold = .96

                  #set up a list

                  detectedlist = [9,9,9,9,9,9,9]
                  if eligible1 >= detectedthreshold:
                      detectedlist[0] = eligibleclass1
                  if eligible2 >= detectedthreshold:
                      detectedlist[1] = eligibleclass2
                  if eligible3 >= detectedthreshold:
                      detectedlist[2] = eligibleclass3
                  if eligible4 >= detectedthreshold:
                      detectedlist[3] = eligibleclass4
                  if eligible5 >= detectedthreshold:
                      detectedlist[4] = eligibleclass5


                  if 1 in detectedlist:
                      if 2 in detectedlist:
                          features[2] = features[2]+1
                      else:
                          features[1] = features[1]+1
                          photolist.append('gable'+str(imgs))
                  if 2 in detectedlist:
                      features[2] = features[2]+1
                      photolist.append('hip'+str(imgs))
                  if 3 in detectedlist:
                      features[3] = features[3]+1
                      photolist.append('chimney'+str(imgs))
                  if 4 in detectedlist:
                      features[4] = features[4]+1
                      photolist.append('vented'+str(imgs))
                  if 5 in detectedlist:
                      features[5] = features[5]+1
                      photolist.append('sat'+str(imgs))
                  if 6 in detectedlist:
                      features[6] = features[6]+1
                      photolist.append('skylight'+str(imgs))
                  if 7 in detectedlist:
                      features[7] = features[7]+1
                      photolist.append('solar'+str(imgs))
                  imgs +=1




            if features[1] > features[2] or features[1]==features[2]:
                roofstyle = 'gable'
                print('This is a gable roof')
                countsome = int(1)
                rawSF = float(0)
                eaves = float(0)
                rakes = float(0)
                valleys = float(0)
                ridges = float(0)
                elements = (len(measurelist))//3
                while countsome <= elements:
                    slope = float(slope)
                    length = float((measurelist[(0+((countsome-1)*3))]))
                    width = float(measurelist[1+((countsome-1)*3)])
                    overlap = float(measurelist[2+((countsome-1)*3)])
                    rafter = float(0)
                    if slope == 0:
                        rafter = .5*width
                    elif slope == 1:
                        angle = 4.76
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 2:
                        angle = 9.46
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 3:
                        angle = 14.04
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 4:
                        angle = 18.43
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 5:
                        angle = 22.62
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 6:
                        angle = 26.57
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 7:
                        angle = 30.26
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 8:
                        angle = 33.69
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 9:
                        angle = ((33.69+39.81)//2)
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 10:
                        angle = 39.81
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 11:
                        angle = 42.51
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 12:
                        angle = 45
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 13:
                        angle = 47.29
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 14:
                        angle = 49.4
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 15:
                        angle = 51.34
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 16:
                        angle = 53.13
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    else:
                        angle = 22.62
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))

                    rawSF+= 2*(float(rafter)*float(length))
                    eaves += 2*float(length)
                    rakes += 2*float(rafter)
                    if countsome >=1:
                        valleys += (float(rafter)*(float(1)+(float(.1)*float(slope))))
                    if countsome == 0:
                        ridges += float(length)
                    if countsome >=1:
                        ridges += float(length+(math.sqrt((valleys*valleys)-(rafter*rafter))))
                        eaves += float(-1*(overlap))
                    countsome+=1





            if features[2] > features[1]:
                roofstyle = 'hip'
                print('This is a hip roof')
                countsome = int(1)
                rawSF = float(0)
                eaves = float(0)
                rakes = float(0)
                valleys = float(0)
                ridges = float(0)
                elements = (len(measurelist))//3
                while countsome <= elements:
                    slope = float(slope)
                    length = float((measurelist[(0+((countsome-1)*3))]))
                    width = float(measurelist[1+((countsome-1)*3)])
                    overlap = float(measurelist[2+((countsome-1)*3)])
                    rafter = float(0)
                    if slope == 0:
                        rafter = .5*width
                    elif slope == 1:
                        angle = 4.76
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 2:
                        angle = 9.46
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 3:
                        angle = 14.04
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 4:
                        angle = 18.43
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 5:
                        angle = 22.62
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 6:
                        angle = 26.57
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 7:
                        angle = 30.26
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 8:
                        angle = 33.69
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 9:
                        angle = ((33.69+39.81)//2)
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 10:
                        angle = 39.81
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 11:
                        angle = 42.51
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 12:
                        angle = 45
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 13:
                        angle = 47.29
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 14:
                        angle = 49.4
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 15:
                        angle = 51.34
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    elif slope == 16:
                        angle = 53.13
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))
                    else:
                        angle = 22.62
                        rafter = ((.5*width)//(math.cos(math.radians(angle))))

                    rawSF+= 2*(.5*float(rafter)*width)+2*(((length+(length-rafter))//2)*rafter)
                    eaves += 2*float(length)+2*float(length)
                    rakes += 0
                    if countsome >=1:
                        valleys += (float(rafter)*(float(1)+(float(.1)*float(slope))))
                    if countsome == 0:
                        ridges += float(length*1.2)
                    if countsome >=1:
                        ridges += float(length+(math.sqrt((valleys*valleys)-(rafter*rafter))))
                        eaves += float(-1*(overlap))
                    countsome+=1

            if features[3] > 0:
                chimneycount = float(1)
                print('Chimeny detected, used default of 1 chimney, manufally verify')
            if features[4] > 0:
                ventedridge = float(1)
                print('Vented ridge detected')
            if features[5] > 0:
                satcount = float(1)
                print('Satellite dish detected, used default of 1 sat dish, manually verify')
            if features[6] > 0:
                skylightcount = float(1)
                print('Skylights present, manually verify count')
            if features[7] > 0:
                solarcount = float(1)
                print('Solar panels present, manually verify count')



            ###############Estimating
            dripedge = eaves+rakes
            sqfoot = float(rawSF)
            squares = float(sqfoot//100)

            #excel

            f=open(PATH_TO_TEST_INFO, "r")
            contentss=f.read()
            lines = contentss.splitlines()
            clientname = lines[1][:-3]
            if len(clientname) <= 1:
                clientname = x
            f.close()
            wb = openpyxl.load_workbook('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
            sheet = wb.get_sheet_by_name('Roofing')

            #drop down

            dvroofing = DataValidation(type="list", formula1='"Laminated Shingles, 3-Tab Shingles, S-Tile, Flat Tile, Metal Roofing, Bit Roofing"', allow_blank=True)
            sheet.add_data_validation(dvroofing)
            rooftotal = float(0)
            roofSummary = []
            if roofstyle == 'hip':
                roofSummary.append('hip')
            else:
                roofSummary.append('gable')

            if rooftype == 1 or rooftype == 2:
                #waste
                waste = float(.1)
                if roofstyle == 'gable':
                    waste = float(.1)
                elif roofstyle == 'hip':
                    waste = float(.15)
                else:
                    roofstyle = 'gable'
                    waste = float(.1)


                #waste
                shingleunitgross = float(squares*(1+waste))
                #rounding for purchaseable quantities
                def x_round(x):
                    return math.ceil(x*3)/3
                shinglequant = x_round(shingleunitgross)
                shinglepurchaseable = round(shinglequant, 2)

                #decision matrix for shingle unit cost
                if laminated > shingle:
                    shingletype = 1
                    roofSummary.append('laminated')
                if shingle > laminated:
                    shingletype = 2
                    roofSummary.append('shingle')
                if shingle == laminated:
                    shingletype = 2
                    roofSummary.append('laminated')

                sheet['A12'] = 'Shingle Removal'
                val0 = sheet['A12']
                dvroofing.add(val0)
                sheet['B12'] = squares
                sheet['C12'] = 'Squares'
                sheet['D12'] = '=vlookup(A12, Pricelist!A1:B100, 2, FALSE)'
                sheet['E12'] = 'Per Square'
                sheet['F12'] = '='
                sheet['G12'] = "=b12*d12"
                rooftotal += float(squares)*float(pricelist.get("Shingle Removal"))
                intList = pricelist2.get("Shingle Removal")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)


                #Roof Felt - excel


                sheet['A14'] = 'Roof Felt'
                val1 = sheet['A14']
                dvroofing.add(val1)
                val1.value = 'Roof Felt'
                sheet['B14'] = squares
                sheet['C14'] = 'Squares'
                sheet['D14'] = '=vlookup(A14, Pricelist!A1:B100, 2, FALSE)'
                sheet['E14'] = 'Per Square'
                sheet['F14'] = ' = '
                sheet['G14'] = '=b14*d14'
                rooftotal += float(squares)*float(pricelist.get("Roof Felt"))
                intList = pricelist2.get("Roof Felt")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)

                if shingletype == 1:

                    sheet['A16'] = 'Laminated Shingles'
                    val2 = sheet['A16']
                    dvroofing.add(val2)
                    val2.value = 'Laminated Shingles'
                    sheet['B16'] = shinglepurchaseable
                    sheet['C16'] = 'Squares'
                    sheet['D16'] = '=vlookup(A16, Pricelist!A1:B100, 2, FALSE)'
                    sheet['E16'] = 'Per square'
                    sheet['F16'] = ' = '
                    sheet['G16'] = '=b16*d16'
                    rooftotal += float(shinglepurchaseable)*float(pricelist.get("Laminated Shingles"))
                    intList = pricelist2.get("Laminated Shingles")
                    intMats = intList[1]
                    roofMaterials += float(shinglepurchaseable)*float(intMats)

                if shingletype == 2:

                    sheet['A16'] = '3-Tab Shingles'
                    val3= sheet['A16']
                    val3.value = '3-Tab Shingles'
                    dvroofing.add(val3)
                    sheet['B16'] = shinglepurchaseable
                    sheet['C16'] = 'Squares'
                    sheet['D16'] = '=vlookup(A16, Pricelist!A1:B100, 2, FALSE)'
                    sheet['E16'] = 'Per square'
                    sheet['F16'] = ' = '
                    sheet['G16'] = '=b16*d14'
                    rooftotal += float(shinglepurchaseable)*float(pricelist.get("3-Tab Shingles"))
                    intList = pricelist2.get("3-Tab Shingles")
                    intMats = intList[1]
                    roofMaterials += float(shinglepurchaseable)*float(intMats)

                #ridges

                ridgeunitcosts = pricelist.get("ridgeunitcost")

                sheet['A18'] = 'Ridge Cap'
                val4 = sheet['A18']
                val4.value = 'Ridge Cap'
                dvroofing.add(val4)
                sheet['B18'] = ridges
                sheet['C18'] = 'ln foot'
                sheet['D18'] = '=vlookup(A18, Pricelist!A1:B100, 2, FALSE)'
                sheet['E18'] = 'per ln foot'
                sheet['F18'] = ' = '
                sheet['G18'] = '=b18*d18'
                rooftotal += float(ridges)*float(pricelist.get("Ridge Cap"))
                intList = pricelist2.get("Ridge Cap")
                intMats = intList[1]
                roofMaterials += float(ridges)*float(intMats)


                #dripedgecost


                dripedge = eaves+rakes

                sheet['A20'] = 'Drip Edge'
                val5 = sheet['A20']
                val5.value = 'Drip Edge'
                dvroofing.add(val5)
                sheet['B20'] = dripedge
                sheet['C20'] = 'ln foot'
                sheet['D20'] = '=vlookup(A20, Pricelist!A1:B100, 2, FALSE)'
                sheet['E20'] = 'per linear foot'
                sheet['F20'] = ' = '
                sheet['G20'] = '=b20*d20'
                rooftotal += float(dripedge)*float(pricelist.get("Drip Edge"))
                intList = pricelist2.get("Drip Edge")
                intMats = intList[1]
                roofMaterials += float(dripedge)*float(intMats)
                #address steep Charges
                z = 2

            #stile roof - excel

            if rooftype == 3:

                roofSummary.append('stile')

                sheet['A12'] = 'Roof Tile Demo'
                val10 = sheet['A12']
                dvroofing.add(val10)
                sheet['B12'] = squares
                sheet['C12'] = 'Squares'
                sheet['D12'] = '=vlookup(A12, Pricelist!A1:B100, 2, FALSE)'
                sheet['E12'] = 'Per Square'
                sheet['F12'] = '='
                sheet['G12'] = "=b12*d12"
                rooftotal += float(squares)*float(pricelist.get("Roof Tile Demo"))
                intList = pricelist2.get("Roof Tile Demo")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)

                #Roof Felt - excel

                sheet['A14'] = 'Roof Felt'
                val1 = sheet['A14']
                dvroofing.add(val1)
                val1.value = 'Roof Felt'
                sheet['B14'] = squares
                sheet['C14'] = 'Squares'
                sheet['D14'] = '=vlookup(A14, Pricelist!A1:B100, 2, FALSE)'
                sheet['E14'] = 'Per Square'
                sheet['F14'] = ' = '
                sheet['G14'] = '=b14*d14'
                rooftotal += float(squares)*float(pricelist.get("Roof Felt"))
                intList = pricelist2.get("Roof Felt")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)



                sheet['A16'] = 'S-Shaped Clay Tiles'
                val12 = sheet['A16']
                dvroofing.add(val12)
                val12.value = 'S-Shaped Clay Tiles'
                sheet['B16'] = squares
                sheet['C16'] = 'Squares'
                sheet['D16'] = '=vlookup(A16, Pricelist!A1:B100, 2, FALSE)'
                sheet['E16'] = 'Per Square'
                sheet['F16'] = ' = '
                sheet['G16'] = '=b16*d16'
                rooftotal += float(squares)*float(pricelist.get("S-Shaped Clay Tiles"))
                intList = pricelist2.get("S-Shaped Clay Tiles")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)
                #ridges

                sheet['A18'] = 'Tile Ridge Cap'
                val13 = sheet['A18']
                dvroofing.add(val13)
                val13.value = 'Tile Ridge Cap'
                sheet['B18'] = ridges
                sheet['C18'] = 'ln foot'
                sheet['D18'] = '=vlookup(A18, Pricelist!A1:B100, 2, FALSE)'
                sheet['E18'] = 'per ln foot'
                sheet['F18'] = ' = '
                sheet['G18'] = '=b18*d18'
                rooftotal += float(ridges)*float(pricelist.get("Tile Ridge Cap"))
                intList = pricelist2.get("Tile Ridge Cap")
                intMats = intList[1]
                roofMaterials += float(ridges)*float(intMats)

                #dripedgecost

                dripedge = eaves+rakes

                sheet['A20'] = 'Drip edge'
                val14 = sheet['A20']
                dvroofing.add(val14)
                val14.value = 'Drip Edge'
                sheet['B20'] = dripedge
                sheet['C20'] = 'ln foot'
                sheet['D20'] = '=vlookup(A20, Pricelist!A1:B100, 2, FALSE)'
                sheet['E20'] = 'per linear foot'
                sheet['F20'] = ' = '
                sheet['G20'] = '=b20*d20'
                rooftotal += float(dripedge)*float(pricelist.get("Drip Edge"))
                intList = pricelist2.get("Drip Edge")
                intMats = intList[1]
                roofMaterials += float(dripedge)*float(intMats)

                #steepcharges

                z = 2

            if rooftype == 4:
                roofSummary.append('metal')

                sheet['A12'] = 'Metal Roofing Removal'
                val40 = sheet['A12']
                dvroofing.add(val40)
                sheet['B12'] = squares
                sheet['C12'] = 'Squares'
                sheet['D12'] = '=VLOOKUP(A12, PriceList!A1:B100, 2, FALSE)'
                sheet['E12'] = 'Per Square'
                sheet['F12'] = '='
                sheet['G12'] = '=b12*d12'
                rooftotal += float(squares)*float(pricelist.get("Metal Roofing Removal"))
                intList = pricelist2.get("Metal Roofing Removal")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)

                #Roof Felt - excel

                sheet['A14'] = 'Roof Felt'
                val1 = sheet['A14']
                dvroofing.add(val1)
                val1.value = 'Roof Felt'
                sheet['B14'] = squares
                sheet['C14'] = 'Squares'
                sheet['D14'] = '=vlookup(A14, Pricelist!A1:B100, 2, FALSE)'
                sheet['E14'] = 'Per Square'
                sheet['F14'] = ' = '
                sheet['G14'] = '=b14*d14'
                rooftotal += float(squares)*float(pricelist.get("Roof Felt"))
                intList = pricelist2.get("Roof Felt")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)

                #roofing material

                sheet['A16'] = 'Metal Roofing'
                val41 = sheet['A16']
                dvroofing.add(val41)
                val41.value = 'Metal Roofing'
                sheet['B16'] = squares
                sheet['C16'] = 'Squares'
                sheet['D16'] = '=vlookup(A16, Pricelist!A1:B100, 2, FALSE)'
                sheet['E16'] = 'Per Square'
                sheet['F16'] = ' = '
                sheet['G16'] = '=b16*d16'
                rooftotal += float(squares)*float(pricelist.get("Metal Roofing"))
                intList = pricelist2.get("Metal Roofing")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)

                #ridges

                sheet['A18'] = 'Metal Ridge Cap'
                val42 = sheet['A18']
                dvroofing.add(val42)
                val42.value = 'Metal Ridge Cap'
                sheet['B18'] = round(ridges)
                sheet['C18'] = 'ln foot'
                sheet['D18'] = '=vlookup(A18, Pricelist!A1:B100, 2, FALSE)'
                sheet['E18'] = 'per ln foot'
                sheet['F18'] = ' = '
                sheet['G18'] = '=b18*d18'
                rooftotal += float(ridges)*float(pricelist.get("Metal Ridge Cap"))
                intList = pricelist2.get("Metal Ridge Cap")
                intMats = intList[1]
                roofMaterials += float(ridges)*float(intMats)

                #dripedgecost

                dripedge = eaves+rakes

                sheet['A20'] = 'Drip Edge'
                val43 = sheet['A20']
                dvroofing.add(val43)
                val43.value = 'Drip Edge'
                sheet['B20'] = dripedge
                sheet['C20'] = 'ln foot'
                sheet['D20'] = '=vlookup(A20, Pricelist!A1:B100, 2, FALSE)'
                sheet['E20'] = 'per linear foot'
                sheet['F20'] = ' = '
                sheet['G20'] = '=b18*d18'
                rooftotal += float(dripedge)*float(pricelist.get("Drip Edge"))
                intList = pricelist2.get("Drip Edge")
                intMats = intList[1]
                roofMaterials += float(dripedge)*float(intMats)

                #steepcharges
                z = 2

            if rooftype == 5:

                roofSummary.append('flat tile')

                sheet['A12'] = 'Roof Tile Demo'
                val40 = sheet['A12']
                dvroofing.add(val40)
                val40.value = 'Roof Tile Demo'
                sheet['B12'] = squares
                sheet['C12'] = 'Squares'
                sheet['D12'] = '=VLOOKUP(A12, PriceList!A1:B100, 2, FALSE)'
                sheet['E12'] = 'Per Square'
                sheet['F12'] = '='
                sheet['G12'] = '=b12*d12'
                rooftotal += float(squares)*float(pricelist.get("Roof Tile Demo"))
                intList = pricelist2.get("Roof Tile Demo")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)
                #Roof Felt - excel

                sheet['A14'] = 'Roof Felt'
                val1 = sheet['A14']
                dvroofing.add(val1)
                val1.value = 'Roof Felt'
                sheet['B14'] = squares
                sheet['C14'] = 'Squares'
                sheet['D14'] = '=vlookup(A14, Pricelist!A1:B100, 2, FALSE)'
                sheet['E14'] = 'Per Square'
                sheet['F14'] = ' = '
                sheet['G14'] = '=b14*d14'
                rooftotal += float(squares)*float(pricelist.get("Roof Felt"))
                intList = pricelist2.get("Roof Felt")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)

                #roofing material

                sheet['A16'] = 'Flat Tile'
                val50 = sheet['A16']
                dvroofing.add(val50)
                val50.value = 'Flat Tile'
                sheet['B16'] = squares
                sheet['C16'] = 'Squares'
                sheet['D16'] = '=vlookup(A16, Pricelist!A1:B100, 2, FALSE)'
                sheet['E16'] = 'Per square'
                sheet['F16'] = ' = '
                sheet['G16'] = '=b16*d16'
                rooftotal += float(squares)*float(pricelist.get("Flat Tile"))
                intList = pricelist2.get("Flat Tile")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)


                #ridges

                sheet['A18'] = 'Tile Ridge Cap'
                val51= sheet['A18']
                dvroofing.add(val51)
                val51.value = 'Tile Ridge Cap'
                sheet['B18'] = ridges
                sheet['C18'] = 'ln foot'
                sheet['D18'] = '=vlookup(A18, Pricelist!A1:B100, 2, FALSE)'
                sheet['E18'] = 'per ln foot'
                sheet['F18'] = ' = '
                sheet['G18'] = '=b18*d18'
                rooftotal += float(ridges)*float(pricelist.get("Tile Ridge Cap"))
                intList = pricelist2.get("Tile Ridge Cap")
                intMats = intList[1]
                roofMaterials += float(ridges)*float(intMats)

                val14 = sheet['A20']
                val14.value = 'Drip Edge'
                dvroofing.add(val14)
                sheet['B20'] = dripedge
                sheet['C20'] = 'ln foot'
                sheet['D20'] = '=vlookup(A20, Pricelist!A1:B100, 2, FALSE)'
                sheet['E20'] = 'per linear foot'
                sheet['F20'] = ' = '
                sheet['G20'] = '=b20*d20'
                rooftotal += float(dripedge)*float(pricelist.get("Drip Edge"))
                intList = pricelist2.get("Drip Edge")
                intMats = intList[1]
                roofMaterials += float(dripedge)*float(intMats)

                z = 2

            if rooftype == 6:
                roofSummary.append('built up')

                sheet['A12'] = 'Built Up Removal'
                val60 = sheet['A12']
                dvroofing.add(val60)
                val60.value = 'Built Up Removal'
                sheet['B12'] = squares
                sheet['C12'] = 'Squares'
                sheet['D12'] = '=VLOOKUP(A12, PriceList!A1:B100, 2, FALSE)'
                sheet['E12'] = 'Per Square'
                sheet['F12'] = '='
                sheet['G12'] = '=d12*b12'
                rooftotal += float(squares)*float(pricelist.get("Built Up Removal"))
                intList = pricelist2.get("Built Up Removal")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)

                #Roof Felt - excel

                sheet['A14'] = 'Roof Felt'
                val11 = sheet['A14']
                dvroofing.add(val11)
                val11.value = 'Roof Felt'
                sheet['B14'] = squares
                sheet['C14'] = 'Squares'
                sheet['D14'] = '=vlookup(A14, Pricelist!A1:B100, 2, FALSE)'
                sheet['E14'] = 'Per Square'
                sheet['F14'] = ' = '
                sheet['G14'] = '=b14*d14'
                rooftotal += float(squares)*float(pricelist.get("Roof Felt"))
                intList = pricelist2.get("Roof Felt")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)

                #roofing material

                sheet['A16'] = 'Built Up Roofing'
                val61 = sheet['A16']
                dvroofing.add(val61)
                val61.value = 'Built Up Roofing'
                sheet['B16'] = squares
                sheet['C16'] = 'Squares'
                sheet['D16'] = '=vlookup(A16, Pricelist!A1:B100, 2, FALSE)'
                sheet['E16'] = 'Per square'
                sheet['F16'] = ' = '
                sheet['G16'] = '=b16*d16'
                rooftotal += float(squares)*float(pricelist.get("Built Up Roofing"))
                intList = pricelist2.get("Built Up Roofing")
                intMats = intList[1]
                roofMaterials += float(squares)*float(intMats)


                val14 = sheet['A18']
                val14.value = 'Drip Edge'
                dvroofing.add(val14)
                sheet['B18'] = dripedge
                sheet['C18'] = 'ln foot'
                sheet['D18'] = '=vlookup(A18, Pricelist!A1:B100, 2, FALSE)'
                sheet['E18'] = 'per linear foot'
                sheet['F18'] = ' = '
                sheet['G18'] = '=b20*d20'
                rooftotal += float(dripedge)*float(pricelist.get("Drip Edge"))
                intList = pricelist2.get("Drip Edge")
                intMats = intList[1]
                roofMaterials += float(dripedge)*float(intMats)

                z = 0

            if rooftype == 6:
                z = 0
            else:
                z = 2




            def add_variables(rooftotal, roofMaterials):
                f=open(PATH_TO_TEST_INFO, "r")
                contents =f.read()
                lines = contents.splitlines()
                clientname = lines[1][:3]
                if len(clientname) <= 1:
                    clientname = x
                clientaddress = lines[6][:-3]
                zip = clientaddress[-5:].strip()
                measurelist = []
                slope = 4
                levels = 1
                for line in lines:
                    if line.startswith('measure'):
                        rawMeasure = line[(len('measure ')):]
                        measurelist = rawMeasure.split(',')
                        for each in measurelist:
                            if not each == "":
                                each = float(each)
                        measurelist = measurelist[:-1]

                    if line.startswith('slope'):
                        slope = line[len('slope '):]
                    if line.startswith('levels'):
                        levels = line[7]

                        print('levelsraw are '+levels)


                pitch = float(slope)//float(12)

                rooftotal = float(rooftotal)
                variableSummary = []
                if rooftype == 6:
                    z = 0
                else:
                    z = 2
                verysteepunitcosts = pricelist.get("verysteepunitcost")
                steepunitcosts = pricelist.get("steepunitcost")
                if pitch >= .75:
                    steepcharge = verysteepunitcosts
                elif pitch >= .5 and pitch < .75:
                    steepcharge = steepunitcosts
                else:
                    steepcharge = 0

                #address steep Charges

                if pitch >= .5 and pitch < .75:
                    sheet['A20'] = 'Steep Charge'
                    val15 = sheet['A20']
                    val15.value = 'Steep Charge'
                    dvroofing.add(val15)
                    sheet['B20'] = squares
                    sheet['C20'] = 'squares'
                    sheet['D20'] = '=vlookup(A20, PriceList!A1:B100, 2, FALSE)'
                    sheet['E20'] = 'per square'
                    sheet['F20'] = ' = '
                    sheet['G20'] = '=b22*d22'
                    rooftotal += float(squares)*float(pricelist.get("Steep Charge"))
                    intList = pricelist2.get("Steep Charge")
                    intMats = intList[1]
                    roofMaterials += float(squares)*float(intMats)
                    variableSummary.append('Steep')
                    z = z+2
                if pitch >= .75:
                    sheet['A20'] = 'Very Steep Charge'
                    val16 = sheet['A20']
                    val16.value = 'Very Steep Charge'
                    sheet['B20'] = squares
                    sheet['C20'] = 'squares'
                    sheet['D20'] = '=vlookup(A22, PriceList!A1:B100, 2, FALSE)'
                    sheet['E20'] = 'per square'
                    sheet['F20'] = ' = '
                    sheet['G20'] = '=b22*d22'
                    rooftotal += float(squares)*float(pricelist.get("Very Steep Charge"))
                    intList = pricelist2.get("Very Steep Charge")
                    intMats = intList[1]
                    roofMaterials += float(squares)*float(intMats)
                    variableSummary.append('Very')
                    z=z+2


                stories = int(levels)
                if stories >= 2:
                    sheet['A'+str(z+20)] = 'High Charge'
                    val7 = sheet['A'+str(z+20)]
                    val7.value = 'High Charge'
                    dvroofing.add(val7)
                    val77 = 'A'+str(z+20)
                    sheet['B'+str(z+20)] = squares
                    sheet['C'+str(z+20)] = 'squares'
                    sheet['D'+str(z+20)] = '=vlookup('+str(val77)+', PriceList!A1:B100, 2, FALSE)'
                    sheet['E'+str(z+20)] = 'per square'
                    sheet['F'+str(z+20)] = ' = '
                    sheet['G'+str(z+20)] = '=b'+str(z+20)+'*d'+str(z+20)
                    rooftotal += float(squares)*float(pricelist.get("High Charge"))
                    intList = pricelist2.get("High Charge")
                    intMats = intList[1]
                    roofMaterials += float(squares)*float(intMats)
                    variableSummary.append('High')
                    z = z+2

                #valley valleymetal
                if valleys > 0 :
                    val8 = sheet['A'+str(z+20)]
                    dvroofing.add(val8)
                    val8.value = 'Valley Metal'
                    val88 = 'A'+str(z+20)
                    sheet['B'+str(z+20)] = valleys
                    sheet['C'+str(z+20)] = 'ln feet'
                    sheet['D'+str(z+20)] = '=vlookup('+str(val88)+', PriceList!A1:B100, 2, FALSE)'
                    sheet['E'+str(z+20)] = 'per foot'
                    sheet['F'+str(z+20)] = ' = '
                    sheet['G'+str(z+20)] = '=b'+str(z+20)+'*d'+str(z+20)
                    rooftotal += float(valleys)*float(pricelist.get("Valley Metal"))
                    intList = pricelist2.get("Valley Metal")
                    intMats = intList[1]
                    roofMaterials += float(valleys)*float(intMats)
                    variableSummary.append('valley '+str(valleys))
                    z = z+2

                if squares > 4:
                    val88 = sheet['A'+str(z+20)]
                    dvroofing.add(val88)
                    val88.value = 'Roof Vent Flashing'
                    val888 = 'A'+str(z+20)
                    sheet['B'+str(z+20)] = round(squares/3)
                    sheet['C'+str(z+20)] = 'each'
                    sheet['D'+str(z+20)] = '=vlookup('+str(val888)+', PriceList!A1:B100, 2, FALSE)'
                    sheet['E'+str(z+20)] = 'each'
                    sheet['F'+str(z+20)] = ' = '
                    sheet['G'+str(z+20)] = '=b'+str(z+20)+'*d'+str(z+20)
                    rooftotal += float(round(squares/4.5))*float(pricelist.get("Roof Vent Flashing"))
                    intList = pricelist2.get("Roof Vent Flashing")
                    intMats = intList[1]
                    roofMaterials += float(round(squares/3))*float(intMats)
                    z = z+2


                if float(chimneycount) > float(0):

                    val9 = sheet['A'+str(z+20)]
                    dvroofing.add(val9)
                    val9.value = 'Chimney Flashing'
                    val99 = 'A'+str(z+20)
                    sheet['B'+str(z+20)] = '1'
                    sheet['C'+str(z+20)] = 'per each'
                    sheet['D'+str(z+20)] = '=VLOOKUP('+str(val99)+', PriceList!A1:B100, 2, FALSE)'
                    sheet['E'+str(z+20)] = 'each'
                    sheet['F'+str(z+20)] = ' = '
                    sheet['G'+str(z+20)] = '=b'+str(z+20)+'*d'+str(z+20)
                    rooftotal += float(1)*float(pricelist.get("Chimney Flashing"))
                    intList = pricelist2.get("Chimney Flashing")
                    intMats = intList[1]
                    roofMaterials += float(intMats)
                    variableSummary.append('Chimney')
                    z = z+2

                if ventedridge > float(0):

                    val10 = sheet['A'+str(z+20)]
                    dvroofing.add(val10)
                    val10.value = "Vented Ridge Cap"
                    val101 = 'A'+str(z+20)
                    sheet['B'+str(z+20)] = ridges
                    sheet['C'+str(z+20)] = 'per linear foot'
                    sheet['D'+str(z+20)] = '=VLOOKUP('+str(val101)+', PriceList!A1:B100, 2, FALSE)'
                    sheet['E'+str(z+20)] = 'each'
                    sheet['F'+str(z+20)] = ' = '
                    sheet['G'+str(z+20)] = '=b'+str(z+20)+'*d'+str(z+20)
                    rooftotal += float(ridges)*float(pricelist.get("Vented Ridge Cap"))
                    intList = pricelist2.get("Vented Ridge Cap")
                    intMats = intList[1]
                    roofMaterials += float(ridges)*float(intMats)
                    variableSummary.append('vented')
                    z = z+2

                if satcount > float(0):

                    val11 = sheet['A'+str(z+20)]
                    dvroofing.add(val11)
                    val11.value = 'Detach Satellite Dish'
                    val111 = 'A'+str(z+20)
                    sheet['B'+str(z+20)] = satcount
                    sheet['C'+str(z+20)] = 'each'
                    sheet['D'+str(z+20)] = '=VLOOKUP('+str(val111)+', PriceList!A1:B100, 2, FALSE)'
                    sheet['E'+str(z+20)] = 'each'
                    sheet['F'+str(z+20)] = ' = '
                    sheet['G'+str(z+20)] = '=b'+str(z+20)+'*d'+str(z+20)
                    rooftotal += float(satcount)*float(pricelist.get("Detach Satellite Dish"))
                    intList = pricelist2.get("Detach Satellite Dish")
                    intMats = intList[1]
                    roofMaterials += float(satcount)*float(intMats)
                    variableSummary.append('Satellite')
                    z=z+2

                if skylightcount> float(0):
                    val12 = sheet['A'+str(z+20)]
                    dvroofing.add(val12)
                    val12.value = 'Skylight Flashing'
                    val121 = 'A'+str(z+20)
                    sheet['B'+str(z+20)] = skylightcount
                    sheet['C'+str(z+20)] = 'each'
                    sheet['D'+str(z+20)] = '=VLOOKUP('+str(val121)+', PriceList!A1:B100, 2, FALSE)'
                    sheet['E'+str(z+20)] = 'each'
                    sheet['F'+str(z+20)] = ' = '
                    sheet['G'+str(z+20)] = '=b'+str(z+20)+'*d'+str(z+20)
                    rooftotal += float(skylightcount)*float(pricelist.get("Skylight Flashing"))
                    intList = pricelist2.get("Skylight Flashing")
                    intMats = intList[1]
                    roofMaterials += float(skylightcount)*float(intMats)
                    variableSummary.append('skylight')
                    z = z+2

                if solarcount> float(0):

                    val13 = sheet['A'+str(z+20)]
                    dvroofing.add(val13)
                    val13.value = 'Detach and Reset Solar Panel'
                    val131 = 'A'+str(z+20)
                    sheet['B'+str(z+20)] = solarcount
                    sheet['C'+str(z+20)] = 'each'
                    sheet['D'+str(z+20)] = '=VLOOKUP('+str(val131)+', PriceList!A1:B100, 2, FALSE)'
                    sheet['E'+str(z+20)] = 'each'
                    sheet['F'+str(z+20)] = ' = '
                    sheet['G'+str(z+20)] = '=b'+str(z+20)+'*d'+str(z+20)
                    rooftotal += float(solarcount)*float(pricelist.get("Detach and Reset Solar Panel"))
                    intList = pricelist2.get("Detach and Reset Solar Panel")
                    intMats = intList[1]
                    roofMaterials += float(solarcount)*float(intMats)
                    variableSummary.append('solar')
                    z = z+2



                if squares <= 12:
                    debristype = "Debris Haul"
                    dumpsterquant = 1
                    debriscost = pricelist.get("Debris Haul")
                    variableSummary.append('Trash Haul')
                if squares >12:
                    dumpsterquant = round(squares/20)
                    debristype = "Dumpster"
                    debriscost = pricelist.get("Dumpster")
                    variableSummary.append('Dumpster')

                val14 = sheet['A'+str(z+20)]
                dvroofing.add(val14)
                val14.value = debristype
                sheet['B'+str(z+20)] = dumpsterquant
                sheet['C'+str(z+20)] = 'each'
                sheet['D'+str(z+20)] = debriscost
                sheet['E'+str(z+20)] = 'per each'
                sheet['F'+str(z+20)] = ' = '
                sheet['G'+str(z+20)] = '=b'+str(z+20)+'*d'+str(z+20)
                rooftotal += float(debriscost)*float(dumpsterquant)


                #Subtotal
                sumint = 'G'+str(z+20)
                sheet['E'+str(z+22)] = 'Subtotal'
                sheet['F'+str(z+22)] = '  = '
                sheet['G'+str(z+22)] = '=sum(G12:'+sumint+')'

                #tax
                taxint = 'G'+str(z+22)
                sheet['E'+str(z+24)] = 'Tax'
                sheet['F'+str(z+24)] = ' = '
                sheet['G'+str(z+24)] = '='+str(roofMaterials)+'*'+str(taxrate)

                totalint1 = 'G'+str(z+22)
                totalint2 = 'G'+str(z+24)
                sheet['E'+str(z+26)] = 'Grand Total'
                sheet['F'+str(z+26)] = ' = '
                sheet['G'+str(z+26)] = '='+totalint1+'+'+totalint2

                variableString = ""
                for each in variableSummary:
                    variableString += str(each)+" "

                f=open(PATH_TO_TEST_INFO, "a")
                f.write('\nxxr '+str(rooftotal))
                f.write('\nxxmr '+str(roofMaterials))
                f.write('\nVariable '+variableString)
                f.close()

            add_variables(rooftotal, roofMaterials)


            roofString = ""
            for each in roofSummary:
                roofString += str(each)+" "

            f=open(PATH_TO_TEST_INFO, "a")
            f.write('\nRoof type: '+roofString)
            f.close()

            namelist = []
            for name in os.listdir(PATH_TO_TEST_IMAGES_DIR):
                namelist.append(name)

            sheet7 = wb.get_sheet_by_name('Photos')
            global namecount
            namecount = 1
            for name in namelist:
                image = Image.open(PATH_TO_TEST_IMAGES_DIR+"image"+str(namecount)+".jpg")
                new_image = image.resize((273, 273))
                new_image.save(PATH_TO_TEST_IMAGES_DIR+"image"+str(namecount)+".jpg")
                img_name = PATH_TO_TEST_IMAGES_DIR+"image"+str(namecount)+".jpg"
                img8 = openpyxl.drawing.image.Image(img_name)
                img8.anchor = 'A'+str(namecount+11+((namecount-1)*16))
                sheet7.add_image(img8)
                caption_row_anchor = str((namecount+11+((namecount-1)*16))+14)
                namecount +=1


            f=open(PATH_TO_TEST_INFO, "a")
            f.write('\nRoof type: '+roofString)
            f.close()

            roof_name1 = photolist[0]
            roof_name0 = str(roof_name1)
            if roof_name0.startswith('['):
                roof_name0 = roof_name0[2:-2]

            caption1 = 'This photo shows a '+roof_name0+' roofing product'
            caption2 = 'A '+roof_name0+' product is seen in this photo'
            caption3 = 'In this photo, a '+roof_name0+' product is seen'
            caption4 = 'This photo shows a '+roof_name0+' roofing product'
            caption5 = 'A '+roof_name0+' product is seen in this photo'
            gable_position = ""
            hip_position = ""
            chimney_position = ""
            vented_position = ""
            sat_position = ""
            skylight_position = ""
            solar_position = ""

            for each in photolist:
                if "gable" in each and roofstyle == 'gable':
                    gable_position += str(each[-1:])
                if "hip" in each and roofstyle == 'hip':
                    hip_position += str(each[-1:])
                if 'chimney' in each:
                    chimney_position += str(each[-1:])
                if 'vented' in each:
                    vented_position += str(each[-1:])
                if "sat" in each:
                    sat_position += str(each[-1:])
                if 'skylight' in each:
                    skylight_position += str(each[-1:])
                if 'solar' in each:
                    solar_position += str(each[-1:])
            if len(gable_position) > 0:
                for each in gable_position:
                    if each == "1":
                        caption1+= ' in a gable roofing system'
                    if each == "2":
                        caption2+= ' in a gable roofing system'
                    if each == "3":
                        caption3+= ' in a gable roofing system'
                    if each =="4":
                        caption4 += ' in a gable roofing system'
                    if each =='5':
                        caption5 += ' in a gable roofing system'
            if len(hip_position) > 0:
                for each in hip_position:
                    if each == "1":
                        caption1+= ' in a hip roofing system'
                    if each == "2":
                        caption2+= ' in a hip roofing system'
                    if each == "3":
                        caption3+= ' in a hip roofing system'
                    if each =="4":
                        caption4 += ' in a hip roofing system'
                    if each =='5':
                        caption5 += ' in a hip roofing system'
            if len(chimney_position) > 0:
                for each in chimney_position:
                    if each == "1":
                        caption1+= '.  This photo also shows a chimney.'
                    if each == "2":
                        caption2+= '.  This photo also shows a chimney.'
                    if each == "3":
                        caption3+= '.  This photo also shows a chimney.'
                    if each =="4":
                        caption4 += '.  This photo also shows a chimney.'
                    if each =='5':
                        caption5 += '.  This photo also shows a chimney.'
            if len(vented_position) > 0:
                for each in vented_position:
                    if each == "1":
                        caption1+= ' This photo also shows a vented ridge cap.'
                    if each == "2":
                        caption2+= ' This photo also shows a vented ridge cap.'
                    if each == "3":
                        caption3+= ' This photo also shows a vented ridge cap.'
                    if each =="4":
                        caption4 += ' This photo also shows a vented ridge cap.'
                    if each =='5':
                        caption5 += ' This photo also shows a vented ridge cap.'
            if len(sat_position) > 0:
                for each in sat_position:
                    if each == "1":
                        caption1+= '.  This photo also shows a satellite dish.'
                    if each == "2":
                        caption2+= '.  This photo also shows a satellite dish.'
                    if each == "3":
                        caption3+= '.  This photo also shows a satellite dish.'
                    if each =="4":
                        caption4 += '.  This photo also shows a satellite dish.'
                    if each =='5':
                        caption5 += '.  This photo also shows a satellite dish.'
            if len(skylight_position) > 0:
                for each in skylight_position:
                    if each == "1":
                        caption1+= '.  This photo also shows a skylight.'
                    if each == "2":
                        caption2+= '.  This photo also shows a skylight.'
                    if each == "3":
                        caption3+= '.  This photo also shows a skylight.'
                    if each =="4":
                        caption4 += '.  This photo also shows a skylight.'
                    if each =='5':
                        caption5 += '.  This photo also shows a skylight.'
            if len(solar_position) > 0:
                for each in solar_position:
                    if each == "1":
                        caption1+= '.  This photo also shows a solar panel.'
                    if each == "2":
                        caption2+= '.  This photo also shows a solar panel.'
                    if each == "3":
                        caption3+= '.  This photo also shows a solar panel.'
                    if each =="4":
                        caption4 += '.  This photo also shows a solar panel.'
                    if each =='5':
                        caption5 += '.  This photo also shows a solar panel.'

            sheet7['A27'] =caption1
            if namecount>2:
                sheet7['A44'] = caption2
            if namecount>3:
                sheet7['A61'] = caption3
            if namecount>4:
                sheet7['A78'] = caption4
            if namecount>5:
                sheet7['A94'] = caption5




            wb.save('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
            wb.close()


#persistence2 is the function which runs following siding input
def siding(batch):
    ############################ siding ###############################
    PATH_TO_CKPT3 = 'D:/models/research/object_detection/exported/siding/frozen_inference_graph.pb'
    PATH_TO_LABELS3 = 'D:/models/research/object_detection/exported/siding/labelmap2.pbtxt'
    NUM_CLASSES3 = 7
    PATH_TO_CKPT9 = 'D:/models/research/object_detection/exported/sidingAncillary/frozen_inference_graph.pb'
    PATH_TO_LABELS9 = 'D:/models/research/object_detection/exported/sidingAncillary/frozen_inference_graph.pb'
    NUM_CLASSES9 = 5
    boxweights = [0,0,0,0,0,0,0,0]
    ancillaryWeights = 0
    areas = 0
    heavyweights = 0
    windowArea = 0
    doorArea = 0
    sidingArea = 0
    sidingMaterials = float(0)
    windowList = []
    doorList = []
    garageList = []
    windowcount = 0
    doorcount = 0
    garagecount = 0
    windowArea = 0
    doorArea = 0
    garageArea = 0

    sidingSummary = []

    sidingraw = 0
    x = batch
    for path in pathlib.Path('D:/models/research/object_detection/server/connect/'+x+'/siding/').iterdir():
        if path.is_file():
            sidingraw = sidingraw+1

    if sidingraw >= 1:
        PATH_TO_TEST_IMAGES_DIR = 'D:/models/research/object_detection/server/connect/'+x+'/siding/'
        PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
                    #rename the photos
        def rename():
            i = 1
            for name in os.listdir(PATH_TO_TEST_IMAGES_DIR):
                dst = PATH_TO_TEST_IMAGES_DIR + "image" + str(i) + ".jpg"
                src = PATH_TO_TEST_IMAGES_DIR + name
                os.rename(src, dst)
                i += 1

            # execute functions

        rename()

        sidingraw+=1
        TEST_IMAGE_PATHS3 = [ os.path.join(PATH_TO_TEST_IMAGES_DIR, 'image{}.jpg'.format(i)) for i in range(1, sidingraw)]

        #simple counters
        board = 0
        brick = 0
        lap = 0
        log = 0
        shingled = 0
        stone = 0
        stucco = 0

        eligibleclass1 = 0
        eligibleclass2 = 0
        eligibleclass3 = 0
        eligibleclass4 = 0
        eligibleclass5 = 0

        area1 = 0
        area2 = 0
        area3 = 0
        area4 = 0
        area5 = 0



        ##########extract the data from PDF into a string
        pdfFileObj = open('eagleview.pdf', 'rb')
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
        pageObj = pdfReader.getPage(0)
        raws = (pageObj.extractText())
        raw = str(raws)

        #define units
        sq = 'sqaures'
        sqft = 'square feet'
        lnft = 'linear foot'
        inc = 'as incurred'

        measurelist = []

        PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
        f=open(PATH_TO_TEST_INFO, "r")
        contents =f.read()
        lines = contents.splitlines()
        levels = 1
        for line in lines:
            if line.startswith('measure'):
                rawMeasure = line[(len('measure ')):]
                measurelist = rawMeasure.split(',')
                for each in measurelist:
                    if not each == "":
                        each = float(each)
                measurelist = measurelist[:-1]
                print(measurelist)
            if line.startswith('slope'):
                slope = line[len('slope '):]


            if line.startswith('levels'):
                levels = line[7]

        print('levelsraw siding are '+str(levels))


        countsome = int(1)
        rawSF = float(0)
        eaves = float(0)
        rakes = float(0)
        ridges = float(0)
        elements = (len(measurelist))//3
        sidingtotal = float(0)
        while countsome <= elements:
            length = float(float(measurelist[(0+((countsome-1)*3))])-float(.75))
            width = float(float(measurelist[1+((countsome-1)*3)])-float(1.5))
            overlap = float(float(measurelist[(2+((countsome-1)*3))])-float(.75))
            print('overlap is '+str(overlap))
            rawSF+= (float(2)*length*(float(levels)*float(12))+(float(2)*width*float(levels)*float(12)))-(overlap*12)

            countsome+=1


        #set up inference siding
        def run_inference_for_single_image(image, sess, tensor_dict):
            if 'detection_masks' in tensor_dict:
                # The following processing is only for single image
                detection_boxes = tf.squeeze(tensor_dict['detection_boxes'], [0])
                detection_masks = tf.squeeze(tensor_dict['detection_masks'], [0])
                # Reframe is required to translate mask from box coordinates to image coordinates and fit the image size.
                real_num_detection = tf.cast(tensor_dict['num_detections'][0], tf.int32)
                detection_boxes = tf.slice(detection_boxes, [0, 0], [real_num_detection, -1])
                detection_masks = tf.slice(detection_masks, [0, 0, 0], [real_num_detection, -1, -1])
                detection_masks_reframed = utils_ops.reframe_box_masks_to_image_masks(
                    detection_masks, detection_boxes, image.shape[0], image.shape[1])
                detection_masks_reframed = tf.cast(
                    tf.greater(detection_masks_reframed, 0.5), tf.uint8)
                # Follow the convention by adding back the batch dimension
                tensor_dict['detection_masks'] = tf.expand_dims(detection_masks_reframed, 0)
            image_tensor = tf.get_default_graph().get_tensor_by_name('image_tensor:0')

            # Run inference
            output_dict = sess.run(tensor_dict, feed_dict={image_tensor: np.expand_dims(image, 0)})

            # all outputs are float32 numpy arrays, so convert types as appropriate
            output_dict['num_detections'] = int(output_dict['num_detections'][0])
            output_dict['detection_classes'] = output_dict[
                'detection_classes'][0].astype(np.uint8)
            output_dict['detection_boxes'] = output_dict['detection_boxes'][0]
            output_dict['detection_scores'] = output_dict['detection_scores'][0]
            if 'detection_masks' in output_dict:
                output_dict['detection_masks'] = output_dict['detection_masks'][0]
            return output_dict

        detection_graph = tf.Graph()
        with detection_graph.as_default():
          od_graph_def = tf.GraphDef()
          with tf.gfile.GFile(PATH_TO_CKPT3, 'rb') as fid:
            serialized_graph = fid.read()
            od_graph_def.ParseFromString(serialized_graph)
            tf.import_graph_def(od_graph_def, name='')

        label_map = label_map_util.load_labelmap(PATH_TO_LABELS3)
        categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES3, use_display_name=True)
        category_index = label_map_util.create_category_index(categories)

        def load_image_into_numpy_array(image):
          (im_width, im_height) = image.size
          return np.array(image.getdata()).reshape(
              (im_height, im_width, 3)).astype(np.uint8)



        with detection_graph.as_default():
          with tf.Session(graph=detection_graph) as sess:
            tf.enable_eager_execution()
            sess.run(tf.global_variables_initializer())
            tf.enable_eager_execution()
            ops = tf.get_default_graph().get_operations()
            tf.enable_eager_execution()
            all_tensor_names = {output.name for op in ops for output in op.outputs}
            tensor_dict = {}
            for key in [
                'num_detections', 'detection_boxes', 'detection_scores',
                'detection_classes', 'detection_masks'
            ]:
                tensor_name = key + ':0'
                if tensor_name in all_tensor_names:
                    tensor_dict[key] = tf.get_default_graph().get_tensor_by_name(tensor_name)
            img = 1
            imgs = 1
            captions1 = ''
            captions2 = ''
            captions3 = ''
            captions4 = ''
            for image_path in TEST_IMAGE_PATHS3:
              tf.enable_eager_execution()
              image = Image.open(image_path)
              image_np = load_image_into_numpy_array(image)
              # Expand dimensions since the model expects images to have shape: [1, None, None, 3]
              image_np_expanded = np.expand_dims(image_np, axis=0)
              image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')
              # Each box represents a part of the image where a particular object was detected.
              boxes = detection_graph.get_tensor_by_name('detection_boxes:0')
              scores = detection_graph.get_tensor_by_name('detection_scores:0')
              classes = detection_graph.get_tensor_by_name('detection_classes:0')
              num_detections = detection_graph.get_tensor_by_name('num_detections:0')
              output_dict = run_inference_for_single_image(image_np, sess, tensor_dict)
              dictionary = output_dict['detection_scores']
              detectedscores = output_dict['detection_scores']
              detectedclasses = output_dict['detection_classes']
              detectedboxes = output_dict['detection_boxes']


              eligible1 = detectedscores[0]
              eligible2 = detectedscores[1]
              eligible3 = detectedscores[2]
              eligible4 = detectedscores[3]
              eligible5 = detectedscores[4]
              eligibleclass1 = detectedclasses[0]
              eligibleclass2 = detectedclasses[1]
              eligibleclass3 = detectedclasses[2]
              eligibleclass4 = detectedclasses[3]
              eligibleclass5 = detectedclasses[4]
              box1 = detectedboxes[0]
              box2 = detectedboxes[1]
              box3 = detectedboxes[2]
              box4 = detectedboxes[3]
              box5 = detectedboxes[4]

             #since this is detection not competition, using lower threshold

              detectionscoreslist = [eligible1, eligible2, eligible3, eligible4, eligible5]


              detectedthreshold = .9

              #set up a list

              detectedlist = [9,9,9,9,9,9,9]
              if eligible1 >= detectedthreshold:
                  detectedlist[0] = eligibleclass1
                  area1 = ((box1[2]-box1[0])*(box1[3]-box1[1]))

              if eligible2 >= detectedthreshold:
                  detectedlist[1] = eligibleclass2
                  area2 = ((box2[2]-box2[0])*(box2[3]-box2[1]))
              if eligible3 >= detectedthreshold:
                  detectedlist[2] = eligibleclass3
                  area3 = ((box3[2]-box3[0])*(box3[3]-box3[1]))
              if eligible4 >= detectedthreshold:
                  detectedlist[3] = eligibleclass4
                  area4 = ((box4[2]-box4[0])*(box4[3]-box4[1]))
              if eligible5 >= detectedthreshold:
                  detectedlist[4] = eligibleclass5
                  area5 = ((box5[2]-box5[0])*(box5[3]-box5[1]))

              if area1 > 0:
                  areas = areas+area1+area2+area3+area4+area5
                  boxweights[eligibleclass1] = (boxweights[eligibleclass1]+area1)
                  boxweights[eligibleclass2] = (boxweights[eligibleclass2]+area2)
                  boxweights[eligibleclass3] = (boxweights[eligibleclass3]+area3)
                  boxweights[eligibleclass4] = (boxweights[eligibleclass4]+area4)
                  boxweights[eligibleclass5] = (boxweights[eligibleclass5]+area5)
              number_siding = 0
              siding_list = []
              for each in detectedlist:
                  if each != 9:
                      number_siding +=1
                      siding_list.append(each)
              photo_string = ""
              siding_type = []
              if 1 in siding_list:
                  siding_type.append('board')
              if 2 in siding_list:
                  siding_type.append('brick')
              if 3 in siding_list:
                  siding_type.append('lap')
              if 4 in siding_list:
                  siding_type.append('log')
              if 5 in siding_list:
                  siding_type.append('shingled')
              if 6 in siding_list:
                  siding_type.append('stone')
              if 7 in siding_list:
                  siding_type.append('stucco')
              zye = len(siding_type)




              if zye == 0:
                  photo_string+= "No siding types detected in this photo."
              if zye ==1:
                  photo_string+= str(siding_type[0]).capitalize()+' siding is seen in this photo'
              if zye ==2:
                  photo_string += str(siding_type[0]).capitalize()+' siding and '+str(siding_type[1])+' siding are seen in this photo.'
              if zye ==3:
                  photo_string+= str(siding_type[0]).capitalize()+' siding, '+str(siding_type[1])+' siding, '+str(siding_type[2])+' siding are seen in this photo'
              if zye >=4:
                  photo_string += str(siding_type[0]).capitalize()+' siding, '+str(siding_type[1])+' siding, '+str(siding_type[2])+' siding are seen in this photo, among others.'


              if imgs == 1:
                 captions1 = photo_string
              if imgs ==2:
                 captions2 = photo_string
              if imgs==3:
                 captions3 =photo_string
              if imgs==4:
                 captions4 =photo_string
              imgs+=1



            #INFERENCE SIDING_ANCILLARY
            def run_inference_for_single_image(image, sess, tensor_dict):
                if 'detection_masks' in tensor_dict:
                    # The following processing is only for single image
                    detection_boxes = tf.squeeze(tensor_dict['detection_boxes'], [0])
                    detection_masks = tf.squeeze(tensor_dict['detection_masks'], [0])
                    # Reframe is required to translate mask from box coordinates to image coordinates and fit the image size.
                    real_num_detection = tf.cast(tensor_dict['num_detections'][0], tf.int32)
                    detection_boxes = tf.slice(detection_boxes, [0, 0], [real_num_detection, -1])
                    detection_masks = tf.slice(detection_masks, [0, 0, 0], [real_num_detection, -1, -1])
                    detection_masks_reframed = utils_ops.reframe_box_masks_to_image_masks(
                        detection_masks, detection_boxes, image.shape[0], image.shape[1])
                    detection_masks_reframed = tf.cast(
                        tf.greater(detection_masks_reframed, 0.5), tf.uint8)
                    # Follow the convention by adding back the batch dimension
                    tensor_dict['detection_masks'] = tf.expand_dims(detection_masks_reframed, 0)
                image_tensor = tf.get_default_graph().get_tensor_by_name('image_tensor:0')

                # Run inference
                output_dict = sess.run(tensor_dict, feed_dict={image_tensor: np.expand_dims(image, 0)})

                # all outputs are float32 numpy arrays, so convert types as appropriate
                output_dict['num_detections'] = int(output_dict['num_detections'][0])
                output_dict['detection_classes'] = output_dict[
                    'detection_classes'][0].astype(np.uint8)
                output_dict['detection_boxes'] = output_dict['detection_boxes'][0]
                output_dict['detection_scores'] = output_dict['detection_scores'][0]
                if 'detection_masks' in output_dict:
                    output_dict['detection_masks'] = output_dict['detection_masks'][0]
                return output_dict

            detection_graph = tf.Graph()
            with detection_graph.as_default():
              od_graph_def = tf.GraphDef()
              with tf.gfile.GFile(PATH_TO_CKPT3, 'rb') as fid:
                serialized_graph = fid.read()
                od_graph_def.ParseFromString(serialized_graph)
                tf.import_graph_def(od_graph_def, name='')

            label_map = label_map_util.load_labelmap(PATH_TO_LABELS3)
            categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES3, use_display_name=True)
            category_index = label_map_util.create_category_index(categories)

            def load_image_into_numpy_array(image):
              (im_width, im_height) = image.size
              return np.array(image.getdata()).reshape(
                  (im_height, im_width, 3)).astype(np.uint8)



            with detection_graph.as_default():
              with tf.Session(graph=detection_graph) as sess:
                tf.enable_eager_execution()
                sess.run(tf.global_variables_initializer())
                tf.enable_eager_execution()
                ops = tf.get_default_graph().get_operations()
                tf.enable_eager_execution()
                all_tensor_names = {output.name for op in ops for output in op.outputs}
                tensor_dict = {}
                for key in [
                    'num_detections', 'detection_boxes', 'detection_scores',
                    'detection_classes', 'detection_masks'
                ]:
                    tensor_name = key + ':0'
                    if tensor_name in all_tensor_names:
                        tensor_dict[key] = tf.get_default_graph().get_tensor_by_name(tensor_name)

                for image_path in TEST_IMAGE_PATHS3:
                  tf.enable_eager_execution()
                  image = Image.open(image_path)
                  image_np = load_image_into_numpy_array(image)
                  # Expand dimensions since the model expects images to have shape: [1, None, None, 3]
                  image_np_expanded = np.expand_dims(image_np, axis=0)
                  image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')
                  # Each box represents a part of the image where a particular object was detected.
                  boxes = detection_graph.get_tensor_by_name('detection_boxes:0')
                  scores = detection_graph.get_tensor_by_name('detection_scores:0')
                  classes = detection_graph.get_tensor_by_name('detection_classes:0')
                  num_detections = detection_graph.get_tensor_by_name('num_detections:0')
                  output_dict = run_inference_for_single_image(image_np, sess, tensor_dict)
                  dictionary = output_dict['detection_scores']
                  detectedscores = output_dict['detection_scores']
                  detectedclasses = output_dict['detection_classes']
                  detectedboxes = output_dict['detection_boxes']
                  ancillaryDict = {"windows":[], "doors":[],"garages":[]}
                  detectedscores = list(detectedscores)
                  detectedclasses = list(detectedclasses)
                  detectedboxes = list(detectedboxes)

                  if 1 in detectedclasses:
                      for each in detectedclasses:
                          if each ==1:
                              windowIndex = detectedclasses.index(each)
                              boxRaw = detectedboxes[windowIndex]
                              area1 = ((boxRaw[2]-boxRaw[0])*(boxRaw[3]-boxRaw[1]))
                              windowList.append(area1)
                  if 2 or 3 in detectedclasses:
                      for each in detectedclasses:
                          if each ==2 or each ==3:
                              doorIndex = detectedclasses.index(each)
                              boxRaw = detectedboxes[doorIndex]
                              area1 = ((boxRaw[2]-boxRaw[0])*(boxRaw[3]-boxRaw[1]))
                              doorList.append(area1)
                  if 4 or 5 in detectedclasses:
                      for each in detectedclasses:
                          if each == 4 or each ==5:
                              garageIndex = detectedclasses.index(each)
                              boxRaw = detectedboxes[garageIndex]
                              area1 = ((boxRaw[2]-boxRaw[0])*(boxRaw[3]-boxRaw[1]))
                              garageList.append(area1)

                ancillaryDict["windows"] = windowList
                ancillaryDict["doors"] = doorList
                ancillaryDict["garage"] = garageList
                windowcount += float(len(ancillaryDict.get('windows')))
                doorcount += float(len((ancillaryDict.get('doors'))))
                garagecount += float(len((ancillaryDict.get('garage'))))
                windowArea += float(sum(ancillaryDict.get('windows')))
                doorArea += float(sum(ancillaryDict.get('doors')))
                garageArea += float(sum(ancillaryDict.get('garage')))
                print('doors are'+str(doorArea))
                print(type(doorArea))
                print(type(windowArea))
                print(type(garageArea))
                print('windows are '+str(windowArea))
                print('garage are '+str(garageArea))
                print('number of windows are '+str(windowcount))
                print('number of doors are '+str(doorcount))
                print('number of garage doors are '+str(garagecount))

                ancillaryWeights += windowArea+doorArea+garageArea

        heavyweights = sum(boxweights)
        ancillarySum = ancillaryWeights

        totalsiding = float(rawSF*(1-(float(ancillarySum)//float(heavyweights))))

        if heavyweights > 0 :
            shareboard = float((boxweights[1]/heavyweights))
            sharebrick = float((boxweights[2]/heavyweights))
            sharelap = float((boxweights[3]/heavyweights))
            sharelog = float((boxweights[4]/heavyweights))
            shareshingled = float((boxweights[5]/heavyweights))
            sharestone = float((boxweights[6]/heavyweights))
            sharestucco = float((boxweights[7]/heavyweights))

            print('The total share of board siding is ',shareboard)
            print('The total share of brick siding is: ',sharebrick)
            print('The total share of lap siding is: ',sharelap)
            print('The total share of log siding is: ',sharelog)
            print('THe total share of shingled siding is: ',shareshingled)
            print('The total share of stone siding is :', sharestone)
            print('THe total area of stucco siding is :', sharestucco)
        else:
            shareboard = float(.00001)
            sharelap = float(.000001)
            sharebrick = float(.000001)
            sharestone = float(.000001)
            sharestone = float(.000001)
            sharestucco = float(.000001)
            shareshingled = float(.00001)
            sharelog = float(.000001)
            print('None detected, used dummy values')

        PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
        f=open(PATH_TO_TEST_INFO, "r")
        contents =f.read()
        lines = contents.splitlines()
        clientname = lines[1][:-3]
        if len(clientname) <=1:
            clientname = x
        clientaddress = lines[6][:-3]
        zip = clientaddress[-5:].strip()


        if zip.isnumeric()==True and len(zip)==5 :
            labor_analysis(zip)
            csv_file = csv.reader(open('zips.csv', "r"), delimiter=",")


            #loop through the csv list
            for row in csv_file:
                #if current rows 2nd value is equal to input, print that row
                if zip == row[0]:
                    taxrate = row[1]


        else:

            taxrate=.07

        f.close()
        wb2 = openpyxl.load_workbook('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
        sheet3 = wb2.get_sheet_by_name('Siding')

        dvsiding = DataValidation(type="list", formula1='"Board and Batten,Masonry Siding,Lap Siding,Log Siding,Shingled Siding,Stone Siding,Stucco,House Wrap"', allow_blank=True)
        sheet3.add_data_validation(dvsiding)

        sheet3['A12'] = 'House Wrap'
        val0 = sheet3['A12']
        dvsiding.add(val0)
        sheet3['B12'] = round(totalsiding)
        sheet3['C12'] = 'sq ft'
        sheet3['D12'] = '=VLOOKUP(A12, PriceList!A1:B100, 2, FALSE)'
        sheet3['E12'] = 'per sq ft'
        sheet3['G12'] = '=b12*d12'
        sidingtotal += round(float(totalsiding))*(float(pricelist.get("House Wrap")))
        sidingList = pricelist2.get("House Wrap")
        sidingMats = sidingList[1]
        sidingMaterials += float(round(totalsiding))*float(sidingMats)


        #counter for cells
        y = 0
        photo_list = []


        if shareboard >= .01:
            sidingSummary.append('board '+str(round(totalsiding*shareboard)))
            val1 = sheet3['A14']
            val1.value = "Board and Batten"
            val12 = 'A14'
            dvsiding.add(val1)
            sheet3['B14'] = round(totalsiding*shareboard)
            sheet3['C14'] = 'sq ft'
            sheet3['D14'] = '=VLOOKUP('+val12+', PriceList!A1:B100, 2, FALSE)'
            sheet3['E14'] = 'per sq ft'
            sheet3['G14'] = '=b14*d14'
            sidingtotal += float(round(totalsiding*shareboard))*float(pricelist.get("Board and Batten"))
            sidingList = pricelist2.get("Board and Batten")
            sidingMats = sidingList[1]
            sidingMaterials += float(round(totalsiding*shareboard))*float(sidingMats)

            y = y+2

        if sharebrick >= .01:
            sidingSummary.append('brick '+str(round(totalsiding*sharebrick)))

            sheet3['A'+str(14+y)] = 'Masonry Siding'
            val2 = sheet3['A'+str(14+y)]
            val21 = 'A'+str(14+y)
            dvsiding.add(val2)
            sheet3['B'+str(14+y)] = round(totalsiding*sharebrick)
            sheet3['C'+str(14+y)] = 'sq ft'
            sheet3['D'+str(14+y)] = '=VLOOKUP('+val21+', PriceList!A1:B100, 2, FALSE)'
            sheet3['E'+str(14+y)] = 'per sq ft'
            sheet3['G'+str(14+y)] = '=B'+str(14+y)+'*d'+str(14+y)
            sidingtotal += float(round(totalsiding*sharebrick))*float(pricelist.get("Masonry Siding"))
            sidingList = pricelist2.get("Masonry Siding")
            sidingMats = sidingList[1]
            sidingMaterials += float(round(totalsiding*sharebrick))*float(sidingMats)

            y = y+2

        if sharelap >= .01:
            sidingSummary.append('lap '+str(round(totalsiding*sharelap)))

            val3 = sheet3['A'+str(14+y)]
            val3.value = "Lap Siding"
            dvsiding.add(val3)
            val31 = 'A'+str(14+y)
            sheet3['B'+str(14+y)] = round(totalsiding*sharelap)
            sheet3['C'+str(14+y)] = 'sq ft'
            sheet3['D'+str(14+y)] = '=VLOOKUP('+val31+', PriceList!A1:B100, 2, FALSE)'
            sheet3['E'+str(14+y)] = 'per sq ft'
            sheet3['G'+str(14+y)] = '=B'+str(14+y)+'*d'+str(14+y)
            sidingtotal += float(round(totalsiding*sharelap))*float(pricelist.get("Lap Siding"))
            sidingList = pricelist2.get("Lap Siding")
            sidingMats = sidingList[1]
            sidingMaterials += float(round(totalsiding*sharelap))*float(sidingMats)
            y = y+2

        if sharelog >= .01:
            sidingSummary.append('log '+str(round(totalsiding*sharelog)))

            val4 = sheet3['A'+str(14+y)]
            val41 = 'A'+str(14+y)
            dvsiding.add(val4)
            val4.value = "Log Siding"
            sheet3['B'+str(14+y)] = round(sharelog*totalsiding)
            sheet3['C'+str(14+y)] = 'sq ft'
            sheet3['D'+str(14+y)] = '=VLOOKUP('+val41+', PriceList!A1:B100, 2, FALSE)'
            sheet3['E'+str(14+y)] = 'per sq ft'
            sheet3['G'+str(14+y)] = '=B'+str(14+y)+'*d'+str(14+y)
            sidingtotal += float(round(totalsiding*sharelog))*float(pricelist.get("Log Siding"))
            sidingList = pricelist2.get("Log Siding")
            sidingMats = sidingList[1]
            sidingMaterials += float(round(totalsiding*sharelog))*float(sidingMats)
            y = y+2

        if shareshingled >= .01:
            sidingSummary.append('shingled '+str(round(totalsiding*shareshingled)))

            val5 = sheet3['A'+str(14+y)]
            val5.value = "Shingled Siding"
            dvsiding.add(val5)
            val51 = 'A'+str(14+y)
            sheet3['B'+str(14+y)] = round(shareshingled*totalsiding)
            sheet3['C'+str(14+y)] = 'sq ft'
            sheet3['D'+str(14+y)] = '=VLOOKUP('+val51+', PriceList!A1:B100, 2, FALSE)'
            sheet3['E'+str(14+y)] = 'per sq ft'
            sheet3['G'+str(14+y)] = '=B'+str(14+y)+'*d'+str(14+y)
            sidingtotal += float(round(totalsiding*shareshingled))*float(pricelist.get("Shingled Siding"))
            sidingList = pricelist2.get("Shingled Siding")
            sidingMats = sidingList[1]
            sidingMaterials += float(round(totalsiding*shareshingled))*float(sidingMats)

            y = y+2

        if sharestone >= .01:
            sidingSummary.append('stone '+str(round(totalsiding*sharestone)))

            val6 = sheet3['A'+str(14+y)]
            val6.value = "Stone Siding"
            dvsiding.add(val6)
            val61 = 'A'+str(14+y)
            sheet3['B'+str(14+y)] = round(sharestone*totalsiding)
            sheet3['C'+str(14+y)] = 'square foot'
            sheet3['D'+str(14+y)] = '=VLOOKUP('+val61+', PriceList!A1:B100, 2, FALSE)'
            sheet3['E'+str(14+y)] = 'per square ft'
            sheet3['G'+str(14+y)] = '=B'+str(14+y)+'*d'+str(14+y)
            sidingtotal += float(round(totalsiding*sharestone))*float(pricelist.get("Stone Siding"))
            sidingList = pricelist2.get("Stone Siding")
            sidingMats = sidingList[1]
            sidingMaterials += float(round(totalsiding*sharestone))*float(sidingMats)

            y = y+2

        if sharestucco >= .01:
            sidingSummary.append('stucco '+str(round(totalsiding*sharestucco)))
            val7 = sheet3['A'+str(14+y)]
            val7.value = "Stucco Siding"
            dvsiding.add(val7)
            val71 = 'A'+str(14+y)
            sheet3['B'+str(14+y)] = round(sharestucco*totalsiding)
            sheet3['C'+str(14+y)] = 'square foot'
            sheet3['D'+str(14+y)] = '=VLOOKUP('+val71+', PriceList!A1:B100, 2, FALSE)'
            sheet3['E'+str(14+y)] = 'per square ft'
            sheet3['G'+str(14+y)] = '=B'+str(14+y)+'*d'+str(14+y)
            sidingtotal += float(round(totalsiding*sharestucco))*float(pricelist.get("Stucco Siding"))
            sidingList = pricelist2.get("Stucco Siding")
            sidingMats = sidingList[1]
            sidingMaterials += float(round(totalsiding*sharestucco))*float(sidingMats)

            y = y+2

        print('levels for siding are '+str(levels))
        if float(levels) >0:
            stories = float(levels)
        else:
            stories = float(1)
        print('stories for siding are '+str(stories))


        if sharestucco >.01 or sharelap >.01 or shareshingled >.01 or shareboard>.01:
            val77 = sheet3['A'+str(14+y)]
            val77.value = "Corner Trim"
            dvsiding.add(val77)
            val771 = 'A'+str(14+y)
            sheet3['B'+str(14+y)] = round(stories*14*4)
            sheet3['C'+str(14+y)] = 'linear foot'
            sheet3['D'+str(14+y)] = '=VLOOKUP('+val771+', PriceList!A1:B100, 2, FALSE)'
            sheet3['E'+str(14+y)] = 'per lineal foot'
            sheet3['G'+str(14+y)] = '=B'+str(14+y)+'*d'+str(14+y)
            sidingtotal += float(round(stories*14*4))*float(pricelist.get("Corner Trim"))
            sidingList = pricelist2.get("Stucco Siding")
            sidingMats = sidingList[1]
            sidingMaterials += float(round(totalsiding*sharestucco))*float(sidingMats)

            y = y+2


        #debriscost
        debristype = 0
        debrishaulunitcost = pricelist.get("Debris Haul")
        dumpsterunitcost = pricelist.get("Dumpster")
        if totalsiding <= 500:
            debriscost = debrishaulunitcost
            debristype = 0
            dumpsterquant = 1
            sidingSummary.append('Debris Haul')
        if totalsiding > 500:
            dumpsterquant = round(totalsiding/1250)
            debriscost = dumpsterquant*dumpsterunitcost
            debirstype = 1
            sidingSummary.append('Dumpsters'+str(dumpsterquant))

        sheet3['A'+str(y+14)] = 'Debris Removal'
        sheet3['B'+str(y+14)] = dumpsterquant
        sheet3['C'+str(y+14)] = 'each'
        sheet3['D'+str(y+14)] = debriscost
        sheet3['E'+str(y+14)] = 'each'
        sheet3['G'+str(14+y)] = '=B'+str(14+y)+'*d'+str(14+y)
        sidingtotal += float(dumpsterquant)*float(debriscost)

        #Subtotal
        sumint = 'G'+str(y+14)
        sheet3['E'+str(y+16)] = 'Subtotal'
        sheet3['F'+str(y+16)] = '  = '
        sheet3['G'+str(y+16)] = '=sum(G12:'+sumint+')'


        #tax
        taxint = 'G'+str(y+16)
        sharematerial = '.4'
        sheet3['E'+str(y+18)] = 'Tax'
        sheet3['F'+str(y+18)] = ' = '
        sheet3['G'+str(y+18)] = '='+str(sidingMaterials)+'*'+str(taxrate)


        #grandtotal
        totalint1 = 'G'+str(y+16)
        totalint2 = 'G'+str(y+18)
        sheet3['E'+str(y+20)] = 'Grand Total'
        sheet3['F'+str(y+20)] = ' = '
        sheet3['G'+str(y+20)] = '='+totalint1+'+'+totalint2

        sidingString = ""
        for each in sidingSummary:
            sidingString += str(each)+" "

        namelist = []
        for name in os.listdir(PATH_TO_TEST_IMAGES_DIR):
            namelist.append(name)
        wb2.save('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
        sheet7 = wb2.get_sheet_by_name('Photos')

        global namecount
        global sidingcount
        if namecount <= 1:
            namecount = 1
        sidingcount = 1
        for name in namelist:
            image = Image.open(PATH_TO_TEST_IMAGES_DIR+"image"+str(sidingcount)+".jpg")
            new_image = image.resize((273, 273))
            new_image.save(PATH_TO_TEST_IMAGES_DIR+"image"+str(sidingcount)+".jpg")
            img_name = PATH_TO_TEST_IMAGES_DIR+"image"+str(sidingcount)+".jpg"
            img8 = openpyxl.drawing.image.Image(img_name)
            img8.anchor = 'A'+str(namecount+11+((namecount-1)*16)+((sidingcount-1)*16))
            sheet7.add_image(img8)
            caption_row_anchor = str((namecount+11+((namecount-1)*16))+14)
            sidingcount +=1



        starting_photo = namecount+11+((namecount-1)*16)
        starting_caption = starting_photo+15

        sheet7['A'+str(starting_caption)] = captions1

        if len(captions2)>0:
            sheet7['A'+str(starting_caption+16)]=captions2
        if len(captions3)>0:
            sheet7['A'+str(starting_caption+32)]=captions3
        if len(captions4)>0:
            sheet7['A'+str(starting_caption+50)]=captions4

        wb2.save('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
        wb2.close()
        f=open(PATH_TO_TEST_INFO, "a")
        f.write('\nxxs '+str(sidingtotal))
        f.write('\nxxms '+str(sidingMaterials))
        f.write('\nSiding '+sidingString)
        f.close()

def fencing(batch):


        fencingraw = 0
        x = batch
        for path in pathlib.Path('D:/models/research/object_detection/server/connect/'+x+'/fencing/').iterdir():
          if path.is_file():
                   fencingraw =  fencingraw+1

        PATH_TO_TEST_IMAGES_DIR = 'D:/models/research/object_detection/server/connect/'+x+'/fencing/'
        PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'

        if  fencingraw >= 1:
            def rename():
              n = 1
              PATH_TO_TEST_IMAGES_DIR = 'D:/models/research/object_detection/server/connect/'+x+'/fencing/'
              PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
              for name in os.listdir(PATH_TO_TEST_IMAGES_DIR):
                  PATH_TO_TEST_IMAGES_DIR = 'D:/models/research/object_detection/server/connect/'+x+'/fencing/'
                  dst = PATH_TO_TEST_IMAGES_DIR + "image" + str(n) + ".jpg"
                  src = PATH_TO_TEST_IMAGES_DIR + name
                  os.rename(src, dst)
                  n = n+1
            rename()


            board = 0
            chainlink = 0
            iron = 0
            picket = 0
            rail = 0
            fenceMaterials = float(0)
            global namecount
            global sidingcount
            namecount = namecount
            sidingcount = sidingcount

            f=open(PATH_TO_TEST_INFO, "r")
            contents =f.read()
            lines = contents.splitlines()
            clientname = lines[1][:-3]
            if len(clientname) <=1:
                clientname = x
            f.close()
            fencingraw+=1
            TEST_IMAGE_PATHS = [ os.path.join(PATH_TO_TEST_IMAGES_DIR, 'image{}.jpg'.format(i)) for i in range(1, fencingraw) ]

            #simple counters

            features = [0,0,0,0,0]
            featuresclasses = ['board', 'chainlink', 'iron', 'picket', 'rail']

            #set up inference
            def run_inference_for_single_image(image, sess, tensor_dict):
                if 'detection_masks' in tensor_dict:
                    # The following processing is only for single image
                    detection_boxes = tf.squeeze(tensor_dict['detection_boxes'], [0])
                    detection_masks = tf.squeeze(tensor_dict['detection_masks'], [0])
                    # Reframe is required to translate mask from box coordinates to image coordinates and fit the image size.
                    real_num_detection = tf.cast(tensor_dict['num_detections'][0], tf.int32)
                    detection_boxes = tf.slice(detection_boxes, [0, 0], [real_num_detection, -1])
                    detection_masks = tf.slice(detection_masks, [0, 0, 0], [real_num_detection, -1, -1])
                    detection_masks_reframed = utils_ops.reframe_box_masks_to_image_masks(
                        detection_masks, detection_boxes, image.shape[0], image.shape[1])
                    detection_masks_reframed = tf.cast(
                        tf.greater(detection_masks_reframed, 0.5), tf.uint8)
                    # Follow the convention by adding back the batch dimension
                    tensor_dict['detection_masks'] = tf.expand_dims(detection_masks_reframed, 0)
                image_tensor = tf.get_default_graph().get_tensor_by_name('image_tensor:0')

                # Run inference
                output_dict = sess.run(tensor_dict, feed_dict={image_tensor: np.expand_dims(image, 0)})

                # all outputs are float32 numpy arrays, so convert types as appropriate
                output_dict['num_detections'] = int(output_dict['num_detections'][0])
                output_dict['detection_classes'] = output_dict[
                    'detection_classes'][0].astype(np.uint8)
                output_dict['detection_boxes'] = output_dict['detection_boxes'][0]
                output_dict['detection_scores'] = output_dict['detection_scores'][0]
                if 'detection_masks' in output_dict:
                    output_dict['detection_masks'] = output_dict['detection_masks'][0]
                return output_dict

            detection_graph = tf.Graph()
            with detection_graph.as_default():
              od_graph_def = tf.GraphDef()
              with tf.gfile.GFile(PATH_TO_CKPT5, 'rb') as fid:
                serialized_graph = fid.read()
                od_graph_def.ParseFromString(serialized_graph)
                tf.import_graph_def(od_graph_def, name='')

            label_map = label_map_util.load_labelmap(PATH_TO_LABELS5)
            categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES, use_display_name=True)
            category_index = label_map_util.create_category_index(categories)

            def load_image_into_numpy_array(image):
              (im_width, im_height) = image.size
              return np.array(image.getdata()).reshape(
                  (im_height, im_width, 3)).astype(np.uint8)



            with detection_graph.as_default():
              with tf.Session(graph=detection_graph) as sess:
                tf.enable_eager_execution()
                sess.run(tf.global_variables_initializer())
                tf.enable_eager_execution()
                ops = tf.get_default_graph().get_operations()
                tf.enable_eager_execution()
                all_tensor_names = {output.name for op in ops for output in op.outputs}
                tensor_dict = {}
                for key in [
                    'num_detections', 'detection_boxes', 'detection_scores',
                    'detection_classes', 'detection_masks'
                ]:
                    tensor_name = key + ':0'
                    if tensor_name in all_tensor_names:
                        tensor_dict[key] = tf.get_default_graph().get_tensor_by_name(tensor_name)
                img = 1
                for image_path in TEST_IMAGE_PATHS:
                  tf.enable_eager_execution()
                  image = Image.open(image_path)
                  image_np = load_image_into_numpy_array(image)
                  # Expand dimensions since the model expects images to have shape: [1, None, None, 3]
                  image_np_expanded = np.expand_dims(image_np, axis=0)
                  image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')
                  # Each box represents a part of the image where a particular object was detected.
                  boxes = detection_graph.get_tensor_by_name('detection_boxes:0')
                  scores = detection_graph.get_tensor_by_name('detection_scores:0')
                  classes = detection_graph.get_tensor_by_name('detection_classes:0')
                  num_detections = detection_graph.get_tensor_by_name('num_detections:0')
                  output_dict = run_inference_for_single_image(image_np, sess, tensor_dict)
                  dictionary = output_dict['detection_scores']
                  best_score = dictionary.item(0)
                  second_best_score = dictionary.item(1)
                  dictionarys = output_dict['detection_classes']
                  best_label = dictionarys.item(0)
                  second_best_label= dictionarys.item(1)

                  #analysis of tesnsorsss
                  #throw out ambiguous
                  if best_score <= .9:
                    print('This photo yielded nothing useful')
                  if best_score >= .9 and second_best_score >= .8:
                    print('This photo is ambiguous, did not count')
                  #verify - labels
                  if best_label == 1 and best_score >= .9:
                    board = board + 1
                    print('I have now seen ', board, ' board fence so far')
                  if best_label == 2 and best_score >= .9:

                    chainlink = chainlink + 1
                    print('I have now seen ', chainlink, ' chain link fences so far')
                  if best_label == 3 and best_score >= .9:
                    iron = iron + 1
                    print('I have no seen ', iron, ' wrought iron fences photos so far')
                  if best_label == 4 and best_score >= .9:
                    picket = picket +1
                    print('I have now seen ', picket ,' picket fences so far')
                  if best_label == 5 and best_score >= .9:
                    rail = rail + 1
                    print('AI determined rail fence.  Totalrail fence = ', rail)


            fencelist = [0, board, chainlink, iron, picket, rail]
            max_count = fencelist.index(max(fencelist))

            fenceSummary = []

            if max_count == 1:
                fencetype = 1
                caption = 'This photo shows a board fence.'
                print('The fence was determinated to be board')
            elif max_count == 2:
                fencetype = 2
                caption = 'This photo shows a chain link fence'
                print('The fence was determined to be chain link')
            elif max_count == 3:
                fencetype = 3
                caption = 'This photo shows an iron fence'
                print('The fence was determined to be iron')
            elif max_count == 4:
                fencetype = 4
                caption = 'This photo shows a picket fence'
                print('The fence type was determinated to be picket')
            elif max_count == 5:
                fencetype = 5
                caption = 'This photo shows a rail fence'
                print('The fence type was determined to be rail')
            else:
                fencetype = 7
                caption = 'This photo shows a fence.'
                print('the fence type was indeterminate')

            #local admin extraction

            PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
            f=open(PATH_TO_TEST_INFO, "r")
            contents =f.read()
            lines = contents.splitlines()
            clientname = lines[1][:-3]
            replace = lines[7][:-3]
            if len(replace) < 1:
                replace = 10
            repaint = lines[8][:-3]
            if len(repaint) <1:
                repaint = 0
            clientaddress = lines[6][:-3]
            zip = clientaddress[-5:].strip()

            if zip.isnumeric()==True and len(zip)==5 :
                labor_analysis(zip)
                #open csv file with zips
                csv_file = csv.reader(open('zips.csv', "r"), delimiter=",")


                #loop through the csv list
                for row in csv_file:
                    #if current rows 2nd value is equal to input, print that row
                    if zip == row[0]:
                        taxrate = row[1]
                        print('taxrate is ',taxrate)

            else:
                print('tax rate not found')
                taxrate=.07

            f.close()

            ###############Estimating

            wb = openpyxl.load_workbook('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
            sheet = wb.get_sheet_by_name('Fencing')
            sheet2 = wb.get_sheet_by_name('PriceList')

            #drop down

            dvfencing = DataValidation(type="list", formula1='"Board Fencing, Chain Link Fence, Wrought Iron, Picket Fence, Rail Fence"', allow_blank=True)
            sheet.add_data_validation(dvfencing)
            fencetotal = float(0)
            y = 0
            if fencetype == 1:

                sheet['A12'] = 'Board Fence'
                val2 = sheet['A12']
                dvfencing.add(val2)
                val2.value = 'Board Fence'
                sheet['B12'] = replace
                sheet['C12'] = 'Ln Ft'
                sheet['D12'] = '=vlookup(A12, Pricelist!A1:B100, 2, FALSE)'
                sheet['E12'] = 'Per Ln ft'
                sheet['F12'] = ' = '
                sheet['G12'] = '=b12*d12'
                fencetotal += float(replace)*float(pricelist.get("Board Fence"))
                fenceList = pricelist2.get("Board Fence")
                fenceMats = fenceList[1]
                fenceMaterials += float(replace)*float(fenceMats)
                fenceSummary.append("Board "+str(replace))

            if fencetype == 2:

                sheet['A12'] = 'Chain Link Fence'
                val2 = sheet['A12']
                dvfencing.add(val2)
                val2.value = 'Chain Link Fence'
                sheet['B12'] = replace
                sheet['C12'] = 'Ln Ft'
                sheet['D12'] = '=vlookup(A12, Pricelist!A1:B100, 2, FALSE)'
                sheet['E12'] = 'Per Ln ft'
                sheet['F12'] = ' = '
                sheet['G12'] = '=b12*d12'
                fencetotal += float(replace)*float(pricelist.get("Chain Link Fence"))
                fenceList = pricelist2.get("Chain Link Fence")
                fenceMats = fenceList[1]
                fenceMaterials += float(replace)*float(fenceMats)
                fenceSummary.append("Chain "+str(replace))


            if fencetype == 3:

                sheet['A12'] = 'Wrought Iron Fence'
                val2 = sheet['A12']
                dvfencing.add(val2)
                val2.value = 'Wrought Iron Fence'
                sheet['B12'] = replace
                sheet['C12'] = 'Ln Ft'
                sheet['D12'] = '=vlookup(A12, Pricelist!A1:B100, 2, FALSE)'
                sheet['E12'] = 'Per Ln ft'
                sheet['F12'] = ' = '
                sheet['G12'] = '=b12*d12'
                fencetotal += float(repaint)*float(pricelist.get("Wrought Iron Fence"))
                fenceSummary.append("Wrought "+str(replace))


            if fencetype == 4:

                sheet['A12'] = 'Picket Fence'
                val2 = sheet['A12']
                dvfencing.add(val2)
                val2.value = 'Picket Fence'
                sheet['B12'] = replace
                sheet['C12'] = 'Ln Ft'
                sheet['D12'] = '=vlookup(A12, Pricelist!A1:B100, 2, FALSE)'
                sheet['E12'] = 'Per Ln ft'
                sheet['F12'] = ' = '
                sheet['G12'] = '=b12*d12'
                fencetotal += float(repaint)*float(pricelist.get("Picket Fence"))
                fenceList = pricelist2.get("Picket Fence")
                fenceMats = fenceList[1]
                fenceMaterials += float(replace)*float(fenceMats)
                fenceSummary.append("Picket "+str(replace))

            if fencetype == 5:

                sheet['A12'] = 'Rail Fence'
                val2 = sheet['A12']
                dvfencing.add(val2)
                val2.value = 'Rail Fence'
                sheet['B12'] = replace
                sheet['C12'] = 'Ln Ft'
                sheet['D12'] = '=vlookup(A12, Pricelist!A1:B100, 2, FALSE)'
                sheet['E12'] = 'Per Ln ft'
                sheet['F12'] = ' = '
                sheet['G12'] = '=b12*d12'
                fencetotal += float(repaint)*float(pricelist.get("Rail Fence"))
                fenceList = pricelist2.get("Rail Fence")
                fenceMats = fenceList[1]
                fenceMaterials += float(replace)*float(fenceMats)
                fenceSummary.append("Rail "+str(replace))

            if float(repaint) > float(0):

                y = 2
                sheet['A14'] = 'Paint Fence'
                val3 = sheet['A14']
                dvfencing.add(val3)
                val3.value = 'Paint Fence'
                sheet['B14'] = repaint
                sheet['C14'] = 'Ln Ft'
                sheet['D14'] = '=vlookup(A14, Pricelist!A1:B100, 2, FALSE)'
                sheet['E14'] = 'Per Ln ft'
                sheet['F14'] = ' = '
                sheet['G14'] = '=b14*d14'
                fencetotal += float(repaint)*float(pricelist.get("Paint Fence"))
                fenceList = pricelist2.get("Chain Link Fence")
                fenceMats = fenceList[1]
                fenceMaterials += float(repaint)*float(fenceMats)
                fenceSummary.append("Paint "+str(repaint))

            #Subtotal
            sumint = 'G'+str(y+12)
            sheet['E'+str(y+14)] = 'Subtotal'
            sheet['F'+str(y+14)] = '  = '
            sheet['G'+str(y+14)] = '=sum(G12:'+sumint+')'


            #tax
            taxint = 'G'+str(y+14)
            sheet['E'+str(y+16)] = 'Tax'
            sheet['F'+str(y+16)] = ' = '
            sheet['G'+str(y+16)] = '='+str(fenceMaterials)+'*'+str(taxrate)

            #grandtotal
            totalint1 = 'G'+str(y+14)
            totalint2 = 'G'+str(y+16)
            sheet['E'+str(y+20)] = 'Grand Total'
            sheet['F'+str(y+20)] = ' = '
            sheet['G'+str(y+20)] = '='+totalint1+'+'+totalint2
            wb.save('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
            wb2 = openpyxl.load_workbook('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
            sheet7 = wb2.get_sheet_by_name('Photos')

            namelist = []
            for name in os.listdir(PATH_TO_TEST_IMAGES_DIR):
                namelist.append(name)
            fencecount = 1
            for name in namelist:
                image = Image.open(PATH_TO_TEST_IMAGES_DIR+"image"+str(fencecount)+".jpg")
                new_image = image.resize((273, 273))
                new_image.save(PATH_TO_TEST_IMAGES_DIR+"image"+str(fencecount)+".jpg")
                img_name = PATH_TO_TEST_IMAGES_DIR+"image"+str(fencecount)+".jpg"
                img8 = openpyxl.drawing.image.Image(img_name)
                img8.anchor = 'A'+str(namecount+13+((namecount-1)*16)+((sidingcount-1)*16)+((fencecount-1)*16))
                sheet7.add_image(img8)
                fencecount +=1


            starting_photo = namecount+13+((namecount-1)*16)+((sidingcount-1)*16)
            starting_caption = starting_photo+15
            sheet7['A'+str(starting_caption)]=caption
            if fencecount ==3:
                sheet7['A'+str(starting_caption+17)] = caption

            wb2.save('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
            wb2.close()

            fenceString=""
            for each in fenceSummary:
                fenceString += str(each)+" "

            f=open(PATH_TO_TEST_INFO, "a")
            f.write('\nxxf '+str(fencetotal))
            f.write('\nxxmf '+str(fenceMaterials))
            f.write('\nfence '+str(fenceString))
            f.close()

#the interior estimating relies on Google Visiion's handwriting analysis module
def interior(batch):
    interiorSummary = []
    x = batch
    interiorraw = 0
    yy = 0
    for path in pathlib.Path('D:/models/research/object_detection/server/connect/'+x+'/interior/').iterdir():
        if path.is_file():
            interiorraw = interiorraw+1

    if interiorraw >= 1:
        PATH_TO_TEST_IMAGES_DIR5 = 'D:/models/research/object_detection/server/connect/'+x+'/interior/'
        PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'

        n = 1
        for name in os.listdir(PATH_TO_TEST_IMAGES_DIR5):
            dst = PATH_TO_TEST_IMAGES_DIR5 + "image" + str(n) + ".jpg"
            src = PATH_TO_TEST_IMAGES_DIR5 + name
            os.rename(src, dst)
            n = n+1
        zz = 1

        interiortotal = float(0)
        interiorMaterials = float(0)
        for badas in range(1 , n):

            client = vision.ImageAnnotatorClient()

            file_name = 'image'+str(zz)+'.jpg'
            image_path = PATH_TO_TEST_IMAGES_DIR5+file_name
            print('reviewing file ', image_path)

            with io.open(image_path, 'rb') as image_file:
                content = image_file.read()

            # construct an iamge instance
            image = vision.types.Image(content=content)

            # annotate Image Response
            response = client.document_text_detection(image=image)  # returns TextAnnotation
            df = pd.DataFrame(columns=['locale', 'description'])

            texts = response.text_annotations
            for text in texts:
                df = df.append(
                    dict(
                        locale=text.locale,
                        description=text.description
                    ),
                    ignore_index=True
                )

            text = (df['description'][0])
            #remove nonasciis
            encoded_string = text.encode("ascii", "ignore")
            text = encoded_string.decode()

            print(text)

            insulationUsed = float(0)
            drywallUsed = float(0)
            paintUsed = float(0)
            baseboardsUsed = float(0)
            carpetUsed = float(0)
            vinylUsed = float(0)
            laminatedUsed = float(0)
            engineeredUsed = float(0)


            namestart = text.find("ROOM NAME")
            nameend = text.find("ROOM WIDTH")
            roomnameraw = text[namestart+len('room name'):nameend]
            roompossible = ['Kitchen', 'Living Room', 'Dining','Dining Room', 'Bathroom', 'Family Room', 'Garage', 'Bedroom', 'Hallway', 'Utility Room', 'Office']
            if len(roomnameraw) > 0:
                for letter in roomnameraw:
                    if letter.isalnum()==False:
                        letter = 0
                roomnamecleaned = difflib.get_close_matches(roomnameraw,roompossible,1,.1)
                roomnamecleaned = str(roomnamecleaned)
                roomname=roomnamecleaned[2:-2]
                interiorSummary.append("rrn "+roomname)

            else:
                roomname = 'Room'
                interiorSummary.append("rrn "+roomname)


            widthstart = text.find("ROOM WIDTH")
            widthend = text.find("ROOM LENGTH")
            width = text[widthstart+len('room width'):widthend].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
            width= float(width)

            length0 = text.find("ROOM LENGTH")
            length1 = text.find("ROOM HEIGHT")
            length = text[length0+len('room length'):length1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
            length = float(length)

            height1 = text.find("DOORS")
            if text[length1+len("room height\n")] < text[height1]:
                height = text[length1+len("room height"):height1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('S','5').replace('z','5').replace('Z','5')
                height = float(height)
            else:
                height = float(8.0)

            door1 = text.find("WINDOWS")
            doors = text[height1+len('doors'):door1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')

            windows1 = text.find("INSULATION")
            windows = text[door1+len('windows'):windows1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')


            insulation1 = text.find("DRYWALL")
            insulation = text[windows1+len('insulation'):insulation1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')

            drywall1 = text.find("PAINT")
            drywall = text[(insulation1+len('drywall')):drywall1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')

            paint1 = text.find("BASEBOARDS")
            paint = text[(drywall1+len('paint')):paint1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')

            baseboards1 = text.find("CARPET")
            baseboards = text[paint1+len('baseboards'):baseboards1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')

            carpet1 = text.find("VINYL")
            carpet = text[baseboards1+len('carpet'):carpet1].strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')

            vinyl1 = text.find("LAMINATED")
            vinyl = text[baseboards1+len('vinyl'):vinyl1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')

            laminated1 = text.find("ENGINEERED")
            laminated = text[vinyl1+len('laminated'):laminated1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')

            engineered1 = text.find("STOP")
            engineered = text[laminated1+len('engineered'):engineered1].strip().strip('\n').strip(' ').replace(' ', '').replace('\n', '').replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')

            if int(width)> 0:
                width = float(width)
            if int(length)> 0:
                length = float(length)
            if int(height)> 0:
                height = float(height)
            if len(doors)>0:
                if doors[0].isnumeric()== True:
                    doors = doors.replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
                    doors = float(doors)
                elif doors[0].isnumeric()== False:
                    doors = float(0)
            else:
                doors = float(0)

            if len(windows)>0:
                if windows[0].isnumeric() == True:
                    windows = windows.replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
                    windows = float(windows)
                elif windows[0].isnumeric() == False:
                    windows = float(0)
            else:
                doors = float(0)

            wallss = float(((2*(width+length))*height)-(24*float(doors))-(12*float(windows)))
            ceilingss = float(width*length)
            boths = float(wallss+ceilingss)
            floorss = float(width*length)
            perimeters = float((2*(width+length))-(3.5*doors))

            if len(insulation) > 0:
                if insulation[0].isnumeric() == True:
                    insulation = insulation.replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
                    insulationUsed = float(insulation)
                    interiorSummary.append('Insulation '+str(insulationUsed))
                else:
                    inspossible = ['Walls','Ceilings','Both']
                    insulationwash = difflib.get_close_matches(insulation,inspossible,1,.001)
                    insulation = str(insulationwash)
                    insulation = insulation[2:-2]
                    if insulation == 'Walls':
                        insulationUsed = wallss
                        interiorSummary.append('Insulation '+str(insulationUsed))
                    if insulation == 'Ceilings':
                        insulationUsed = ceilingss
                        interiorSummary.append('Insulation '+str(insulationUsed))
                    if insulation == 'Both':
                        insulationUsed = wallss+ceilingss
                        interiorSummary.append('Insulation '+str(insulationUsed))

            if len(drywall) > 0:
                if drywall[0].isnumeric() == True:
                    drywall = drywall.replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
                    drywallUsed = float(drywall)
                    interiorSummary.append('Drywall '+str(drywallUsed))
                else:
                    inspossible = ['Walls','Ceilings','Both']
                    drywallwash = difflib.get_close_matches(drywall,inspossible,1,.001)
                    drywall = str(drywallwash)
                    drywall = drywall[2:-2]
                    if drywall == 'Walls':
                        drywallUsed = wallss
                        interiorSummary.append('Drywall '+str(drywallUsed))
                    if drywall == 'Ceilings':
                        drywallUsed = ceilingss
                        interiorSummary.append('Drywall '+str(drywallUsed))
                    if drywall == 'Both':
                        drywallUsed = wallss+ceilingss
                        interiorSummary.append('Drywall '+str(drywallUsed))
            if len(paint) > 0:
                if paint[0].isnumeric() == True:
                    paint = paint.replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
                    paintUsed = float(paint)
                    interiorSummary.append('Paint '+str(paintUsed))
                if paint[0].isnumeric() == False:
                    paintpossible = ['Walls','Ceilings','Both']
                    paintwash = str(difflib.get_close_matches(paint,paintpossible,1,.001))
                    paintwash = paintwash[2:-2]
                    if paintwash == 'Walls':
                        paintUsed = float(wallss)
                        interiorSummary.append('Paint '+str(paintUsed))
                    if paintwash == 'Ceilings':
                        paintUsed = float(ceilingss)
                        interiorSummary.append('Paint '+str(paintUsed))
                    if paintwash == 'Both':
                        paintUsed = float(boths)
                        interiorSummary.append('Paint '+str(paintUsed))

            if len(baseboards) > 0:
                if baseboards[0].isnumeric() == True:
                    baseboards = baseboards.replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
                    baseboardsUsed = float(baseboards)
                    interiorSummary.append('Baseboards '+str(baseboardsUsed))
                else:
                    baseboardspossible = ['Perimeter']
                    baseboardswash = str(difflib.get_close_matches(baseboards,baseboardspossible,1,.001))
                    baseboards = baseboardswash[2:-2]
                    baseboards = baseboards
                    if baseboards == 'Perimeter':
                        baseboardsUsed = perimeters
                        interiorSummary.append('Baseboards '+str(baseboardsUsed))

            if int(len(carpet)) > 0:
                if carpet[0].isnumeric() == True:
                    carpet = carpet.replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
                    carpetUsed = float(carpet*1.15)
                    interiorSummary.append('Carpet '+str(carpetUsed))
                else:
                    carpetpossible = ['Floors']
                    carpetwash = difflib.get_close_matches(carpet,carpetpossible,1,.001)
                    carpet = str(carpetwash)
                    carpet = carpet[2:-2]
                    if carpet == 'Floors':
                        carpetUsed = float(floorss)*float(1.15)
                        interiorSummary.append('Carpet '+str(carpetUsed))

            if len(vinyl) > 0:
                if vinyl[0].isnumeric() == True:
                    vinyl = vinyl.replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
                    vinylUsed = float(vinyl)
                    interiorSummary.append('Vinyl '+str(vinylUsed))
                else:
                    vinylpossible = ['Floors']
                    vinylwash = difflib.get_close_matches(vinyl,vinylpossible,1,.001)
                    vinyl = str(vinylwash)
                    vinyl = vinyl[2:-2]
                    if vinyl == 'Floors':
                        vinylUsed = float(floorss)*float(1.15)
                        interiorSummary.append('Vinyl '+str(vinylUsed))



            if len(laminated) > 0:
                if laminated[0].isnumeric() == True:
                    laminated = laminated.replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
                    laminatedUsed = float(laminated)
                    interiorSummary.append('Laminated '+str(laminatedUsed))
                else:
                    laminatedpossible = ['Floors']
                    laminatedwash = difflib.get_close_matches(laminated,laminatedpossible,1,.001)
                    laminated = str(laminatedwash)
                    laminated = laminated[2:-2]
                    if laminated == 'Floors':
                        laminatedUsed = floorss
                        interiorSummary.append('Laminated '+str(laminatedUsed))

            if len(engineered) > 0:
                if engineered[0].isnumeric() == True:
                    engineered = engineered.replace('o', '0').replace('O', '0').replace('s', '5').replace('z','5').replace('Z','5').replace('S','5')
                    engineeredUsed = float(engineered)
                    interiorSummary.append('Engineered '+str(laminatedUsed))
                else:
                    engineeredpossible = ['Walls','Ceilings','Both']
                    engineeredwash = difflib.get_close_matches(engineered,engineeredpossible,1,.001)
                    engineered = str(engineeredwash)
                    engineered = engineered[2:-2]
                    if engineered == 'Floors':
                        engineeredUsed = floorss
                        interiorSummary.append('Engineered '+str(laminatedUsed))

            PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
            f=open(PATH_TO_TEST_INFO, "r")
            contents =f.read()
            lines = contents.splitlines()
            clientname = lines[1][:-3]
            clientaddress = lines[6][:-3]
            zip = clientaddress[-5:].strip()

            if zip.isnumeric()==True and len(zip)==5 :
                labor_analysis(zip)
                csv_file = csv.reader(open('zips.csv', "r"), delimiter=",")


                #loop through the csv list
                for row in csv_file:
                    #if current rows 2nd value is equal to input, print that row
                    if zip == row[0]:
                        taxrate = row[1]
                        print('taxrate is ',taxrate)

            else:
                taxrate=.07
            wb = openpyxl.load_workbook('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
            sheet6 = wb.get_sheet_by_name('Interior')
            dvinterior = DataValidation(type="list", formula1='"Insulation,Paint,Baseboards,Carpet,Vinyl,Laminated,Engineered"', allow_blank=True)
            sheet6.add_data_validation(dvinterior)

            sheet6['A'+str(12+yy)] = roomname


            handy = 0
            flooring = 0

            if insulationUsed > float(0):

                sheet6['A'+str(yy+14)] = 'Insulation'
                val66 = sheet6['A'+str(yy+14)]
                val66.value = 'Insulation'
                dvinterior.add(val66)
                val661 = 'A'+str(yy+14)
                sheet6['B'+str(yy+14)] = insulationUsed
                sheet6['C'+str(yy+14)] = 'each'
                sheet6['D'+str(yy+14)] = '=VLOOKUP('+str(val661)+', Pricelist!A1:B100, 2, FALSE)'
                sheet6['E'+str(yy+14)] = 'each'
                sheet6['G'+str(14+yy)] = '=B'+str(14+yy)+'*d'+str(14+yy)
                yy = yy + 2
                interiortotal += float(insulationUsed)*float(pricelist.get("Insulation"))
                handy = 1
                intList = pricelist2.get("Insulation")
                intMats = intList[1]
                interiorMaterials += float(insulationUsed)*float(intMats)


            if drywallUsed > float(0):

                sheet6['A'+str(yy+14)] = 'Drywall'
                val66 = sheet6['A'+str(yy+14)]
                val66.value = 'Drywall'
                dvinterior.add(val66)
                val661 = 'A'+str(yy+14)
                sheet6['B'+str(yy+14)] = drywallUsed
                sheet6['C'+str(yy+14)] = 'each'
                sheet6['D'+str(yy+14)] = '=VLOOKUP('+str(val661)+', Pricelist!A1:B100, 2, FALSE)'
                sheet6['E'+str(yy+14)] = 'each'
                sheet6['G'+str(14+yy)] = '=B'+str(14+yy)+'*d'+str(14+yy)
                yy = yy + 2
                interiortotal += float(drywallUsed)*float(pricelist.get("Drywall"))
                handy = 1
                intList = pricelist2.get("Drywall")
                intMats = intList[1]
                interiorMaterials += float(drywallUsed)*float(intMats)

            if paintUsed > float(0):

                sheet6['A'+str(yy+14)] = 'Paint'
                val66 = sheet6['A'+str(yy+14)]
                val66.value = 'Paint'
                dvinterior.add(val66)
                val661 = 'A'+str(yy+14)
                sheet6['B'+str(yy+14)] = paintUsed
                sheet6['C'+str(yy+14)] = 'each'
                sheet6['D'+str(yy+14)] = '=VLOOKUP('+str(val661)+', Pricelist!A1:B100, 2, FALSE)'
                sheet6['E'+str(yy+14)] = 'each'
                sheet6['G'+str(14+yy)] = '=B'+str(14+yy)+'*d'+str(14+yy)
                yy = yy + 2
                interiortotal += float(paintUsed)*float(pricelist.get("Paint"))
                handy = 1
                intList = pricelist2.get("Paint")
                intMats = intList[1]
                interiorMaterials += float(paintUsed)*float(intMats)

            if baseboardsUsed > float(0):

                sheet6['A'+str(yy+14)] = 'Baseboards'
                val66 = sheet6['A'+str(yy+14)]
                val66.value = 'Baseboards'
                dvinterior.add(val66)
                val661 = 'A'+str(yy+14)
                sheet6['B'+str(yy+14)] = baseboardsUsed
                sheet6['C'+str(yy+14)] = 'each'
                sheet6['D'+str(yy+14)] = '=VLOOKUP('+str(val661)+', Pricelist!A1:B100, 2, FALSE)'
                sheet6['E'+str(yy+14)] = 'each'
                sheet6['G'+str(14+yy)] = '=B'+str(14+yy)+'*d'+str(14+yy)
                yy = yy + 2
                interiortotal += float(baseboardsUsed)*float(pricelist.get("Baseboards"))
                handy = 1
                intList = pricelist2.get("Baseboards")
                intMats = intList[1]
                interiorMaterials += float(baseboardsUsed)*float(intMats)

            if carpetUsed > float(0):

                sheet6['A'+str(yy+14)] = 'Carpet'
                val66 = sheet6['A'+str(yy+14)]
                val66.value = 'Carpet'
                dvinterior.add(val66)
                val661 = 'A'+str(yy+14)
                sheet6['B'+str(yy+14)] = carpetUsed
                sheet6['C'+str(yy+14)] = 'each'
                sheet6['D'+str(yy+14)] = '=VLOOKUP('+str(val661)+', Pricelist!A1:B100, 2, FALSE)'
                sheet6['E'+str(yy+14)] = 'each'
                sheet6['G'+str(14+yy)] = '=B'+str(14+yy)+'*d'+str(14+yy)
                yy = yy + 2
                interiortotal += float(carpetUsed)*float(pricelist.get("Carpet"))
                handy = 1
                intList = pricelist2.get("Carpet")
                intMats = intList[1]
                interiorMaterials += float(carpetUsed)*float(intMats)

            if vinylUsed > float(0):

                sheet6['A'+str(yy+14)] = 'Vinyl'
                val66 = sheet6['A'+str(yy+14)]
                val66.value = 'Vinyl'
                dvinterior.add(val66)
                val661 = 'A'+str(yy+14)
                sheet6['B'+str(yy+14)] = vinylUsed
                sheet6['C'+str(yy+14)] = 'ach'
                sheet6['D'+str(yy+14)] = '=VLOOKUP('+str(val661)+', Pricelist!A1:B100, 2, FALSE)'
                sheet6['E'+str(yy+14)] = 'each'
                sheet6['G'+str(14+yy)] = '=B'+str(14+yy)+'*d'+str(14+yy)
                yy = yy + 2
                interiortotal += float(vinylUsed)*float(pricelist.get("Vinyl"))
                flooring = 1
                intList = pricelist2.get("Vinyl")
                intMats = intList[1]
                interiorMaterials += float(vinylUsed)*float(intMats)

            if laminatedUsed > float(0):

                sheet6['A'+str(yy+14)] = 'Laminated'
                val66 = sheet6['A'+str(yy+14)]
                val66.value = 'Laminated'
                dvinterior.add(val66)
                val661 = 'A'+str(yy+14)
                sheet6['B'+str(yy+14)] = laminatedUsed
                sheet6['C'+str(yy+14)] = 'each'
                sheet6['D'+str(yy+14)] = '=VLOOKUP('+str(val661)+', Pricelist!A1:B100, 2, FALSE)'
                sheet6['E'+str(yy+14)] = 'each'
                sheet6['G'+str(14+yy)] = '=B'+str(14+yy)+'*d'+str(14+yy)
                yy = yy + 2
                interiortotal += float(laminatedUsed)*float(pricelist.get("Laminated"))
                flooring = 1
                intList = pricelist2.get("Laminated")
                intMats = intList[1]
                interiorMaterials += float(laminatedUsed)*float(intMats)

            if engineeredUsed > float(0) or carpetUsed > float(0) or vinylUsed > float(0):
                reducerUsed = float(doors)*float(4)
                sheet6['A'+str(yy+14)] = 'Reducer Strip'
                val662 = sheet6['A'+str(yy+14)]
                val662.value = 'Reducer Strip'
                dvinterior.add(val662)
                val662 = 'A'+str(yy+14)
                sheet6['B'+str(yy+14)] = reducerUsed
                sheet6['C'+str(yy+14)] = 'each'
                sheet6['D'+str(yy+14)] = '=VLOOKUP('+str(val662)+', Pricelist!A1:B100, 2, FALSE)'
                sheet6['E'+str(yy+14)] = 'each'
                sheet6['G'+str(14+yy)] = '=B'+str(14+yy)+'*d'+str(14+yy)
                yy = yy + 2
                interiortotal += float(reducerUsed)*float(pricelist.get("Reducer Strip"))
                intList = pricelist2.get("Reducer Strip")
                intMats = intList[1]
                interiorMaterials += float(reducerUsed)*float(intMats)


            #add for next ROOM
            yy = yy+2
            zz = zz+1
            wb.save('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
            wb.close()



        #Subtotal

            yy = yy-2
            wb = openpyxl.load_workbook('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
            sheet6 = wb.get_sheet_by_name('Interior')
            sumint = 'G'+str(yy+12)
            sheet6['E'+str(yy+14)] = 'Subtotal'
            sheet6['F'+str(yy+14)] = '  = '
            sheet6['G'+str(yy+14)] = '=sum(G12:'+sumint+')'


            #tax
            taxint = 'G'+str(yy+14)
            sharematerial = '.4'
            sheet6['E'+str(yy+16)] = 'Tax'
            sheet6['F'+str(yy+16)] = ' = '
            sheet6['G'+str(yy+16)] = '='+str(interiorMaterials)+'*'+str(taxrate)



            #grandtotal
            totalint1 = 'G'+str(yy+14)
            totalint2 = 'G'+str(yy+16)
            sheet6['E'+str(yy+18)] = 'Interior Total'
            sheet6['F'+str(yy+18)] = ' = '
            sheet6['G'+str(yy+18)] = '='+totalint1+'+'+totalint2

            wb.save('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
            wb.close()
            trades = int(flooring)+int(handy)
            interiorString = ""
            for each in interiorSummary:
                interiorString+=str(each)+" "

            f=open(PATH_TO_TEST_INFO, "a")
            f.write('\nxxi '+str(trades)+' xxn '+str(interiortotal))
            f.write('\nxxmi '+str(interiorMaterials))
            f.write('\ninterior '+interiorString)
            f.close()

def summary(batch):
        x = batch
        PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
        f=open(PATH_TO_TEST_INFO, "r")
        contents =f.read()
        lines = contents.splitlines()
        clientname = lines[1][:-3]
        clientaddress = lines[6][:-3]
        zip = clientaddress[-5:].strip()
        trades = float(0)
        jj = 0
        roof_summary = 0
        siding_summary = 0
        fence_summary = 0
        interior_summary =0
        total_materials = float(0)

        if zip.isnumeric()==True and len(zip)==5 :

            csv_file = csv.reader(open('zips.csv', "r"), delimiter=",")
            for row in csv_file:
                if zip == row[0]:
                    taxrate = row[1]
                    print('taxrate is ',taxrate)

        else:
            taxrate=.07

        for line in lines:
            if line.startswith('xxi'):
                interior_summary = float(line[10:])
                trades = trades+float(line[4])
                print(interior_summary)
                print('trades are ', trades)
            if line.startswith('xxr'):
                roof_summary = float(line[4:])
                print(roof_summary)
                trades=trades+1
            if line.startswith('xxs'):
                siding_summary = float(line[4:])
                trades=trades+1
            if line.startswith('xxf'):
                trades = trades+1
                fence_summary = float(line[4:])
            if line.startswith('xxmi'):
                total_materials += float(line[5:])
            if line.startswith('xxmr'):
                total_materials += float(line[5:])
            if line.startswith('xxms'):
                total_materials += float(line[5:])
            if line.startswith('xxmf'):
                total_materials += float(line[5:])

        wb = openpyxl.load_workbook('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
        sheet = wb.get_sheet_by_name('Summary')

        sub_summary = 0

        if roof_summary > 0:
            sheet['C12'] = 'Roofing'
            sheet['G12'] = '='+str(roof_summary)

            jj +=2
            sub_summary+=roof_summary

        if interior_summary > 0:
            sheet['C'+str(jj+12)] = 'Interior'
            sheet['G'+str(jj+12)] = '='+str(interior_summary)
            jj += 2
            sub_summary+=interior_summary

        if fence_summary > 0:
            sheet['C'+str(jj+12)] = 'Fencing'
            sheet['G'+str(jj+12)] = '='+str(fence_summary)
            jj += 2
            sub_summary+=fence_summary

        if siding_summary > 0:

            sheet['C'+str(jj+12)] = 'Siding'
            sheet['G'+str(jj+12)] = '='+str(siding_summary)
            jj += 2
            sub_summary+=siding_summary

        sheet['C'+str(jj+12)] = 'Subtotal'
        sheet['G'+str(jj+12)] = '=sum(G12:G'+str(jj+10)+')'
        jj += 2

        sheet['C'+str(jj+12)] = 'Tax'
        sheet['G'+str(jj+12)] = '='+str(total_materials)+'*'+str(taxrate)
        jj += 2

        if trades >= 3 and sub_summary > 10000:
            sheet['C'+str(jj+12)] = 'Overhead and Profit'
            sheet['G'+str(jj+12)] = '=.2*(G'+str(jj+8)+'-G'+str(jj+10)+')'
            jj += 2
            sheet['C'+str(jj+12)] = 'Grand Total'
            sheet['G'+str(jj+12)] = '=sum(G'+str(jj+6)+':G'+str(jj+10)+')'
        else:
            sheet['C'+str(jj+12)] = 'Total'
            sheet['G'+str(jj+12)] = '=sum(G'+str(jj+8)+':G'+str(jj+10)+')'






        wb.save('D:/models/research/object_detection/server/connect/'+x+'/'+clientname+'.xlsx')
        wb.close()

def narrative(batch):
        x = batch
        PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
        f=open(PATH_TO_TEST_INFO, "r")
        contents =f.read()
        lines = contents.splitlines()
        clientname = lines[1][:-3]
        clientaddress = lines[6][:-3]
        zip = clientaddress[-5:].strip()
        trades = float(0)
        jj = 0
        roof_summary = 0
        siding_summary = 0
        fence_summary = 0
        interior_summary =0
        total_materials = float(0)

        if zip.isnumeric()==True and len(zip)==5 :
            csv_file = csv.reader(open('zips.csv', "r"), delimiter=",")
            for row in csv_file:
                if zip == row[0]:
                    taxrate = row[1]
                    print('taxrate is ',taxrate)

        else:
            taxrate=.07

        narrativeString = ""
        for line in lines:
            if line.startswith('xxi'):
                interior_summary = float(line[10:])
                trades = trades+float(line[4])
                print(interior_summary)
                print('trades are ', trades)
            if line.startswith('xxr'):
                roof_summary = float(line[4:])
                print(roof_summary)
                trades=trades+1
            if line.startswith('xxs'):
                siding_summary = float(line[4:])
                trades=trades+1
            if line.startswith('xxf'):
                trades = trades+1
                fence_summary = float(line[4:])
            if line.startswith('xxmi'):
                total_materials += float(line[5:])
            if line.startswith('xxmr'):
                total_materials += float(line[5:])
            if line.startswith('xxms'):
                total_materials += float(line[5:])
            if line.startswith('xxmf'):
                total_materials += float(line[5:])
            if line.startswith('Variable'):
                variableList = line.split()
                narrativeString += "This estimate includes the cost to remove and replace the roofing surface."
                if 'valley' in variableList:
                    narrativeString += '  As valleys were present on the roof, valley metal was added to the estimate.'
                if 'Chimney' in variableList:
                    narrativeString += '  A chimney was noted on the roof, and additional allowances for chimney flashings were made.'
                if 'vented' in variableList:
                    narrativeString += '  Vented ridge cap was noted, and the estimate includes amounts for the replacement of vented rige cap.'
                if 'satellite' in variableList:
                    narrativeString += '  A satellite dish was present on the roof, and allowances were made to detach and reset the satellite dish.'
                if 'skylight' in variableList:
                    narrativeString += '  Skylights were noted on the roofing surface, and we made allowances for skylight flashings.'
                if 'solar' in variableList:
                    narrativeString+= '  Solar panels are presnet on the roof, and the estimate includes amounts for the cost to detach and reset the solar panels.'
                if 'Steep' in variableList:
                    narrativeString += '  As the pitch of the roof is between 6/12 and 9/12, steep charges were added for additional time necessary for safety on a steep roof.'
                if 'Very' in variableList:
                    narrativeString += '  As the pitch of the roof is greater than 9/12, allowances were made for additional safety time for a very steep roof.'
                if 'High' in variableList:
                    narrativeString += '  As there are 2 or more stories on the roof, additional high charge allowances were made for product delivery and safety considerations.'
            if line.startswith('Roof type:'):
                roof_list = line.split()
                roof_shape = roof_list[2]
                roof_product = roof_list[3]
                narrativeString+= '  The roof material is '+roof_product+', and the roof shape is '+roof_shape+'.'
                if roof_product=="shingle" or roof_product=='laminated':
                    if roof_shape=='hip':
                        narrativeString+='  As this was a '+roof_product+' product in a hip roof system, we applied 15% waste.'
                    if roof_shape=='gable':
                        narrativeString+='  As this was a '+roof_product+' product in a gable roof system, we applied 10% waste.'

            if line.startswith('Siding'):
                sidingList = line.split()
                print(sidingList)
                if len(narrativeString) > 0:
                    narrativeString+= "\n\n"
                if 'board' in sidingList:
                    narrativeString+= '  Board siding was found in the siding at this property, and allowances were made to remove and replace the board siding.'
                if 'brick' in sidingList:
                    narrativeString += '  Brick siding was found in the siding at this location, and allowances for a mason to remove and replace the masonry were made.'
                if 'lap' in sidingList:
                    narrativeString += '  Lap siding was noted at the property.  As such, allowances for lap siding were included in the estimate.'
                if 'log' in sidingList:
                    narrativeString+= '  Log siding is present at this location, and allowances were made to remove and replace the log siding.'
                if 'shingled' in sidingList:
                    narrativeString += '  Shingled siding was present at this location, and we included amounts to remove and replace the shingled siding.'
                if 'stone' in sidingList:
                    narrativeString += '  Stone siding was observed at this location, and we included amounts for a stone mason to remove and replace the siding.'
                if 'stucco' in sidingList:
                    narrativeString += '  Stucco was noted at this property, and amounts to remove and replace stucco were included in the estimate.'
                if len(sidingList) > 0:
                    narrativeString+= '  Standard moisture-protective house wrap was included in the estimate, and allowances for garage, corner, window, and door trim were added to the estimate.'

            if line.startswith('fence'):
                fenceList = line.split()
                if len(narrativeString)>0:
                    narrativeString+='\n\n'
                if 'Board' in fenceList:
                    narrativeString+='  Board fencing was noted at the property, and is included in the estimate.'
                if 'Chain' in fenceList:
                    narrativeString+='  The cost to remove and replace the chain link fencing at this location is included in the estimate.'
                if 'Wrought' in fenceList:
                    narrativeString += '  The cost to remove and replace wrought iron fencing at this location is included in the estimate.'
                if 'Picket' in fenceList:
                    narrativeString+= '  The cost to remove and replace the picket fences noted at the property are included in the estimate.'
                if 'Paint' in fenceList:
                    narrativeString+= '  The cost to repaint the fence was also included in the estimate.'
                narrativeString+='Allowances for the replacement of fences are included in the estimate.'

            if line.startswith('interior'):
                interiorString = line
                interiorList = interiorString.split('rrn')
                if len(narrativeString) >0:
                    narrativeString+="\n\n"

                fop = 0
                for each in interiorList:
                    each = str(each)
                    if fop >2:
                        narrativeString+="\n\n"
                    if "interior" in each:
                        asjdl=1
                    elif 'Garage' in each:
                        narrativeString +="In the garage, we noted damages that are included in this estimate."
                    elif 'Living' in each:
                        narrativeString+="In the living room, damages were observed which ware included in this estimate"
                    elif 'Dining' in each:
                        narrativeString+= "Damages were also observed in the dining room."
                    elif 'Kitchen' in each:
                        narrativeString+= 'In the kitchen, we observed damages.'
                    elif 'Bathroom' in each:
                        narrativeString+='In the bathroom, there were damages.'
                    elif 'Family' in each:
                        narrativeString+= "This estimte also includes amounts for repairs in the family room."
                    elif 'Bedroom' in each:
                        narrativeString += 'In the bedroom, we noted damages.'
                    elif 'Hallway' in each:
                        narrativeString += 'Damages were observed in the hallway.'
                    elif 'Utility' in each:
                        narrativeString += 'In the utility room, we noted damages.'
                    elif 'Office' in each:
                        narrativeString += 'Damages were also noted in the office.'
                    else:
                        narrativeString += 'Damages were also noted in another room.'

                    if 'Insulation' in each:
                        narrativeString += "  We have included amounts in the estimate for the replacement of insulation."
                    if 'Drywall' in each:
                        narrativeString += "  The replacement of drywall is also recommended in this room."
                    if 'Paint' in each:
                        narrativeString+= "  The room will need to be repainted, and we have allowed to seal and repaint the room."
                    if 'Baseboards' in each:
                        narrativeString+= "  Baseboards will also need to be replaced in this room, and we have estimated for their inclusion."
                    if 'Carpet' in each:
                        narrativeString+= "  We recommend the replacement of carpet in this room.  We have added for the replacement of carpet and 15% waste."
                    if 'Vinyl' in each:
                        narrativeString+="  The vinyl in this room will also need to be replaced; the estimate includes vinyl replacement, with 15% waste."
                    if 'Laminated' in each:
                        narrativeString+="  The laminated flooring in this room will need to be replaced, and we have included amounts for the replacmenet of laminated flooring in this room."
                    if 'Engineered' in each:
                        narrativeString+= "  We observed damages to the engineered flooring in this room, and have estimated for the replacement of the damaged flooring product."
                    fop +=2


        subSummary = float(0)
        subSummary += float(roof_summary)+float(siding_summary)+float(interior_summary)+float(fence_summary)
        f.close()

        if trades > 3 and subSummary > float(10000):
            overheadAndProfit = True
        else:
            overheadAndProfit = False

        if overheadAndProfit ==True:
            trades = str(trades)
            narrativeString += '\n\nDue to the presence of '+trades[0]+' trades and an amount of work reasonably suggesting the involvement of a general contractor, we have added overhead and profit to the estimate.  The amount of overhead and profit is 20%.'

        print(narrativeString)

"""
        yag = yagmail.SMTP(user='ben@esimtaix.ai', password='Bruce78!')
        for x in batch:
             PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
             f=open(PATH_TO_TEST_INFO, "r")
             contents =f.read()
             lines = contents.splitlines()
             clientname = lines[1][:-3]
             if len(clientname) < 1:
                 clientname = x
             f.close()
             recipientemail = 'benkennedy5386@gmail.com'
             attach = 'D:\\models\\research\\object_detection\\server\\connect\\'+x+'\\'+clientname+'.xlsx'
             body = narrativeString
             subject = 'Please find attached your estimate for '+clientname
             yag.send(to = recipientemail, subject = subject, contents = [body, attach])
             print('emailed file client '+clientname+' to '+recipientemail)
"""



def email_project(batch):
     x= batch
     yag = yagmail.SMTP(user='ben5386simplifyai@gmail.com', password='Bruce78!')
     PATH_TO_TEST_INFO = 'D:/models/research/object_detection/server/connect/'+x+'/'+x+'.txt'
     f=open(PATH_TO_TEST_INFO, "r")
     contents =f.read()
     lines = contents.splitlines()
     clientname = lines[1][:-3]
     if len(clientname) < 1:
         clientname = x
     f.close()
     recipientemail = 'benkennedy5386@gmail.com'
     attach = 'D:\\models\\research\\object_detection\\server\\connect\\'+x+'\\'+clientname+'.xlsx'
     body = 'Estimatix.AI Report for '+clientname
     subject = 'Please find attached your estimate for '+clientname
     yag.send(to = recipientemail, subject = subject, contents = [body, attach])
     print('emailed file client '+clientname+' to '+recipientemail)

#cleans out server for fresh batch, moves to storage
def sort_files(batch):
        x = batch
        src = 'D:/models/research/object_detection/server/connect/'+x+'/'
        dst = 'D:/models/research/object_detection/sent/'+x+'/'
        copy_tree(src, dst)
        shutil.rmtree(src, ignore_errors=True)
        if os.path.isdir(src)==True:
            os.rmdir(src)

#reallyPersistent is the recusrive function that activates.
def reallyPersistent():

        download_info('myfirstapplication-16270.appspot.com')
        batch = get_cookies('D:/models/research/object_detection/server/connect/')

        if len(batch) > 0:
            z = 1
            y = len(batch)
            while z <= y:
                a = z-1
                m = batch[(a)]
                get_admin(m)
                slopey(m)
                measurements(m)
                roofing(m)
                siding(m)
                fencing(m)
                interior(m)
                summary(m)
                narrative(m)
                #email_project(m)
                sort_files(m)
                z+=1
        print('tray empty, sleeping for 60s')
        sleep(60)
        reallyPersistent()

reallyPersistent()
